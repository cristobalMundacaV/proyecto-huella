"""Importadores estructurados para factores y actividades de lote."""

from __future__ import annotations

import csv
import io
import logging
import re
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from django.core.cache import cache
from django.db import IntegrityError, transaction
from django.db.models import Q
from openpyxl import load_workbook

from ..factores import classify_factor, format_activity_display_name, normalize_activity_key
from ..models import (
    EmisionLote,
    Empresa,
    EspecieMadera,
    FactorEmision,
    HistorialCambioLote,
    Lote,
    UnidadOperativa,
)


logger = logging.getLogger(__name__)

FACTOR_IMPORT_CACHE_PREFIX = "factor_import_batch:"
LOTE_IMPORT_CACHE_PREFIX = "lote_import_batch:"
ACTIVITY_IMPORT_CACHE_PREFIX = "activity_import_batch:"
UNIT_IMPORT_CACHE_PREFIX = "unit_import_batch:"
COMPANY_IMPORT_CACHE_PREFIX = "company_import_batch:"
FACTOR_IMPORT_CACHE_TTL_SECONDS = 60 * 60
FACTOR_IMPORT_MAX_BYTES = 5 * 1024 * 1024
ALLOWED_FACTOR_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
MIN_YEAR = 1900
MAX_YEAR = date.today().year + 1

REQUIRED_FACTOR_COLUMNS = ["actividad", "unidad", "factor_emision", "fuente", "anio"]
OPTIONAL_FACTOR_COLUMNS = ["categoria", "actividad_key", "descripcion"]
REQUIRED_LOTE_COLUMNS = [
    "id_lote",
    "empresa",
    "fecha",
    "especie",
    "volumen_m3",
    "origen",
]
REQUIRED_ACTIVITY_COLUMNS = ["actividad", "cantidad", "unidad", "fecha"]
REQUIRED_COMPANY_COLUMNS = [
    "empresa_id",
    "nombre",
    "rut",
    "region",
    "comuna",
    "direccion",
    "rubro",
    "email",
    "telefono",
    "contacto",
    "observaciones",
]
REQUIRED_UNIT_COLUMNS = ["unidad_id", "empresa_id", "nombre", "tipo"]
OPTIONAL_COMPANY_COLUMNS: list[str] = []
OPTIONAL_LOTE_COLUMNS = [
    "empresa_id",
    "unidad_id",
    "tipo_producto",
    "densidad_kg_m3",
    "porcentaje_carbono",
    "estado",
    "observaciones",
]
OPTIONAL_UNIT_COLUMNS = ["region", "comuna", "direccion", "descripcion", "activa"]
TENANT_MISMATCH_WARNING = (
    "empresa_id del archivo difiere de la empresa activa; se importara usando la empresa activa"
)


@dataclass
class ParsedFactorRow:
    row_number: int
    raw: dict
    normalized: dict
    errors: list[str]
    status: str
    is_duplicate: bool = False
    exists_in_db: bool = False
    db_action: str | None = None


@dataclass
class ParsedLoteRow:
    row_number: int
    raw: dict
    normalized: dict
    errors: list[str]
    warnings: list[str]
    status: str
    is_duplicate: bool = False
    exists_in_db: bool = False
    db_action: str | None = None


def _normalize_text(value, *, lower: bool = True) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text.lower() if lower else text


def _parse_decimal(value) -> Decimal:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        raise ValueError("factor_emision es obligatorio")

    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("factor_emision debe ser un numero valido") from exc

    if number <= 0:
        raise ValueError("factor_emision debe ser mayor que cero")

    return number


def _parse_year(value) -> int:
    text = str(value or "").strip()
    if not text:
        raise ValueError("anio es obligatorio")

    try:
        year = int(float(text))
    except (TypeError, ValueError) as exc:
        raise ValueError("anio debe ser un entero valido") from exc

    if year < MIN_YEAR or year > MAX_YEAR:
        raise ValueError(f"anio fuera de rango ({MIN_YEAR}-{MAX_YEAR})")

    return year


def _validate_columns(headers: list[str]) -> list[str]:
    normalized = [_normalize_header_key(header) for header in headers]
    missing = [column for column in REQUIRED_FACTOR_COLUMNS if column not in normalized]
    return missing


def _validate_required_columns(headers: list[str], required_columns: list[str]) -> list[str]:
    normalized = [_normalize_header_key(header) for header in headers]
    return [column for column in required_columns if column not in normalized]


def _normalize_header_key(value) -> str:
    key = _normalize_text(value, lower=True)
    key = unicodedata.normalize("NFD", key)
    key = "".join(char for char in key if unicodedata.category(char) != "Mn")
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    if key in {"año", "aã±o"}:
        return "anio"
    aliases = {
        "ano": "anio",
        "a_o": "anio",
        "a_n": "anio",
        "id_empresa": "empresa_id",
        "id_de_empresa": "empresa_id",
        "id_unidad": "unidad_id",
        "id_unidad_operativa": "unidad_id",
        "lote_id": "id_lote",
        "factor_de_emision": "factor_emision",
        "factor_de_emisi_n": "factor_emision",
        "emision_factor": "factor_emision",
        "direcci_n": "direccion",
        "regi_n": "region",
        "volumen_m": "volumen_m3",
        "volumen_m_3": "volumen_m3",
    }
    return aliases.get(key, key)


def _read_csv_rows(uploaded_file) -> tuple[list[dict], list[str]]:
    raw_bytes = uploaded_file.read()
    if len(raw_bytes) > FACTOR_IMPORT_MAX_BYTES:
        raise ValueError("El archivo supera el tamano maximo permitido")

    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    raw_headers = reader.fieldnames or []
    headers = [_normalize_header_key(header) for header in raw_headers]
    missing = _validate_columns(raw_headers)
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    rows = []
    for index, row in enumerate(reader, start=2):
        normalized_row = {}
        for original_key, value in row.items():
            key = _normalize_header_key(original_key)
            if key:
                normalized_row[key] = value
        rows.append({"row_number": index, **normalized_row})
    return rows, headers


def _read_xlsx_rows(uploaded_file) -> tuple[list[dict], list[str]]:
    raw_bytes = uploaded_file.read()
    if len(raw_bytes) > FACTOR_IMPORT_MAX_BYTES:
        raise ValueError("El archivo supera el tamano maximo permitido")

    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [_normalize_header_key(value) if value is not None else "" for value in next(rows_iter, [])]
    missing = _validate_columns(headers)
    if missing:
        workbook.close()
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    rows = []
    try:
        for row_number, values in enumerate(rows_iter, start=2):
            normalized_row = {}
            values_list = list(values or [])
            for index, header in enumerate(headers):
                if not header:
                    continue
                normalized_row[header] = values_list[index] if index < len(values_list) else None
            rows.append({"row_number": row_number, **normalized_row})
    finally:
        workbook.close()

    return rows, headers


def _read_csv_rows_for_columns(uploaded_file, required_columns: list[str]) -> tuple[list[dict], list[str]]:
    raw_bytes = uploaded_file.read()
    if len(raw_bytes) > FACTOR_IMPORT_MAX_BYTES:
        raise ValueError("El archivo supera el tamano maximo permitido")

    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    raw_headers = reader.fieldnames or []
    headers = [_normalize_header_key(header) for header in raw_headers]
    missing = _validate_required_columns(raw_headers, required_columns)
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    rows = []
    for index, row in enumerate(reader, start=2):
        normalized_row = {}
        for original_key, value in row.items():
            key = _normalize_header_key(original_key)
            if key:
                normalized_row[key] = value
        rows.append({"row_number": index, **normalized_row})
    return rows, headers


def _read_xlsx_rows_for_columns(uploaded_file, required_columns: list[str]) -> tuple[list[dict], list[str]]:
    raw_bytes = uploaded_file.read()
    if len(raw_bytes) > FACTOR_IMPORT_MAX_BYTES:
        raise ValueError("El archivo supera el tamano maximo permitido")

    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [_normalize_header_key(value) if value is not None else "" for value in next(rows_iter, [])]
    missing = _validate_required_columns(headers, required_columns)
    if missing:
        workbook.close()
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    rows = []
    try:
        for row_number, values in enumerate(rows_iter, start=2):
            normalized_row = {}
            values_list = list(values or [])
            for index, header in enumerate(headers):
                if not header:
                    continue
                normalized_row[header] = values_list[index] if index < len(values_list) else None
            rows.append({"row_number": row_number, **normalized_row})
    finally:
        workbook.close()

    return rows, headers


def read_uploaded_factor_rows(uploaded_file) -> tuple[list[dict], str]:
    name = Path(uploaded_file.name or "").suffix.lower()
    if name not in ALLOWED_FACTOR_IMPORT_EXTENSIONS:
        raise ValueError("Solo se permiten archivos CSV o XLSX")

    if name == ".csv":
        rows, _ = _read_csv_rows(uploaded_file)
        return rows, "csv"

    rows, _ = _read_xlsx_rows(uploaded_file)
    return rows, "xlsx"


def _peek_lote_headers_csv(uploaded_file) -> list[str]:
    """Peek at CSV headers without reading the entire file."""
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)  # Reset for subsequent reads
    
    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("latin-1")
    
    reader = csv.DictReader(io.StringIO(text))
    return [_normalize_header_key(header) for header in (reader.fieldnames or [])]


def _peek_lote_headers_xlsx(uploaded_file) -> list[str]:
    """Peek at XLSX headers without reading the entire file."""
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)  # Reset for subsequent reads
    
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, [])
        return [_normalize_header_key(value) if value is not None else "" for value in headers]
    finally:
        workbook.close()


def _get_lote_required_columns(headers: list[str]) -> list[str]:
    """Determine required lote columns based on presence of unidad_id in headers."""
    required = list(REQUIRED_LOTE_COLUMNS)
    
    # If unidad_id is present, empresa becomes optional (can be deduced from unidad_operativa)
    if "unidad_id" in headers and "empresa" in required:
        required.remove("empresa")
    
    return required


def read_uploaded_lote_rows(uploaded_file) -> tuple[list[dict], str]:
    name = Path(uploaded_file.name or "").suffix.lower()
    if name not in ALLOWED_FACTOR_IMPORT_EXTENSIONS:
        raise ValueError("Solo se permiten archivos CSV o XLSX")

    # Peek at headers to determine which columns are required
    if name == ".csv":
        headers = _peek_lote_headers_csv(uploaded_file)
    else:
        headers = _peek_lote_headers_xlsx(uploaded_file)
    
    required_columns = _get_lote_required_columns(headers)

    # Now read with the appropriate required columns
    if name == ".csv":
        rows, _ = _read_csv_rows_for_columns(uploaded_file, required_columns)
        return rows, "csv"

    rows, _ = _read_xlsx_rows_for_columns(uploaded_file, required_columns)
    return rows, "xlsx"


def read_uploaded_unit_rows(uploaded_file, empresa_activa=None) -> tuple[list[dict], str]:
    name = Path(uploaded_file.name or "").suffix.lower()
    if name not in ALLOWED_FACTOR_IMPORT_EXTENSIONS:
        raise ValueError("Solo se permiten archivos CSV o XLSX")

    required_columns = ["unidad_id", "nombre", "tipo"] if empresa_activa else REQUIRED_UNIT_COLUMNS

    if name == ".csv":
        rows, _ = _read_csv_rows_for_columns(uploaded_file, required_columns)
        return rows, "csv"

    rows, _ = _read_xlsx_rows_for_columns(uploaded_file, required_columns)
    return rows, "xlsx"


def read_uploaded_company_rows(uploaded_file) -> tuple[list[dict], str]:
    name = Path(uploaded_file.name or "").suffix.lower()
    if name not in ALLOWED_FACTOR_IMPORT_EXTENSIONS:
        raise ValueError("Solo se permiten archivos CSV o XLSX")

    if name == ".csv":
        rows, _ = _read_csv_rows_for_columns(uploaded_file, REQUIRED_COMPANY_COLUMNS)
        return rows, "csv"

    rows, _ = _read_xlsx_rows_for_columns(uploaded_file, REQUIRED_COMPANY_COLUMNS)
    return rows, "xlsx"


def read_uploaded_activity_rows(uploaded_file) -> tuple[list[dict], str]:
    name = Path(uploaded_file.name or "").suffix.lower()
    if name not in ALLOWED_FACTOR_IMPORT_EXTENSIONS:
        raise ValueError("Solo se permiten archivos CSV o XLSX")

    if name == ".csv":
        rows, _ = _read_csv_rows_for_columns(uploaded_file, REQUIRED_ACTIVITY_COLUMNS)
        return rows, "csv"

    rows, _ = _read_xlsx_rows_for_columns(uploaded_file, REQUIRED_ACTIVITY_COLUMNS)
    return rows, "xlsx"


def _build_row_payload(raw_row: dict) -> tuple[dict, list[str]]:
    errors = []
    normalized = {}

    try:
        normalized["actividad"] = format_activity_display_name(
            _normalize_text(raw_row.get("actividad"), lower=False)
        )
        if not normalized["actividad"]:
            errors.append("actividad es obligatoria")
    except Exception:
        errors.append("actividad invalida")

    try:
        normalized["unidad"] = _normalize_text(raw_row.get("unidad"), lower=False)
        if not normalized["unidad"]:
            errors.append("unidad es obligatoria")
    except Exception:
        errors.append("unidad invalida")

    try:
        normalized["factor_emision"] = _parse_decimal(raw_row.get("factor_emision"))
    except Exception as exc:
        errors.append(str(exc))

    fuente = _normalize_text(raw_row.get("fuente"), lower=False)
    if not fuente:
        errors.append("fuente es obligatoria")
    normalized["fuente"] = fuente

    try:
        normalized["anio"] = _parse_year(raw_row.get("anio") or raw_row.get("año"))
    except Exception as exc:
        errors.append(str(exc))

    classification = classify_factor(
        {
            **raw_row,
            "actividad": normalized.get("actividad"),
            "unidad": normalized.get("unidad"),
            "fuente": normalized.get("fuente"),
        }
    )
    normalized.update(classification)

    return normalized, errors


def _row_lookup_key(normalized_row: dict) -> tuple:
    return (
        normalized_row.get("actividad_key") or normalize_activity_key(normalized_row.get("actividad")),
        _normalize_text(normalized_row.get("unidad"), lower=True),
        str(normalized_row.get("factor_emision")),
        normalized_row.get("anio"),
    )


def _is_readable_activity_name(value: str) -> bool:
    return "_" not in str(value or "")


def _parse_lote_decimal(value, field_name: str) -> Decimal | None:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return None

    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} debe ser un numero valido") from exc

    if number <= 0:
        raise ValueError(f"{field_name} debe ser mayor que cero")

    return number


def _parse_lote_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = str(value or "").strip()
    if not text:
        raise ValueError("fecha es obligatoria")

    formats = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d")
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue

    raise ValueError("fecha no tiene un formato valido")


def _stringify_payload(payload: dict) -> dict:
    result = {}
    for key, value in payload.items():
        if isinstance(value, (date, datetime, Decimal)):
            result[key] = str(value)
        elif hasattr(value, "_meta"):
            result[key] = str(value)
        else:
            result[key] = value
    return result


def _lote_belongs_to_empresa(lote: Lote, empresa: Empresa) -> bool:
    if not lote or not empresa:
        return False

    return (
        lote.empresa_id == empresa.id
        or (
            lote.unidad_operativa is not None
            and lote.unidad_operativa.empresa_id == empresa.id
        )
    )


def _find_lote_in_empresa_scope(id_lote: str, empresa: Empresa):
    if not id_lote or not empresa:
        return None

    return (
        Lote.objects.select_related("empresa", "unidad_operativa", "unidad_operativa__empresa")
        .filter(id_lote=id_lote)
        .filter(Q(empresa=empresa) | Q(unidad_operativa__empresa=empresa))
        .first()
    )


def _find_unit_in_empresa_scope(identifier: str, empresa: Empresa | None = None):
    if not identifier:
        return None

    queryset = UnidadOperativa.objects.select_related("empresa")
    if empresa is not None:
        queryset = queryset.filter(empresa=empresa)

    unit = queryset.filter(unidad_id=identifier).first()
    if unit is not None:
        return unit

    return queryset.filter(nombre__iexact=identifier).first()


def _build_lote_payload(raw_row: dict, empresa_activa=None) -> tuple[dict, list[str], list[str]]:
    errors = []
    warnings = []
    normalized = {
        "id_lote": _normalize_text(raw_row.get("id_lote"), lower=False).upper(),
        "empresa_id": _normalize_text(raw_row.get("empresa_id"), lower=False).upper(),
        "unidad_id": _normalize_text(raw_row.get("unidad_id"), lower=False).upper(),
        "empresa_aserradero": _normalize_text(raw_row.get("empresa"), lower=False),
        "especie": _normalize_text(raw_row.get("especie"), lower=True),
        "origen": _normalize_text(raw_row.get("origen"), lower=False),
        "tipo_producto": _normalize_text(raw_row.get("tipo_producto"), lower=False),
        "estado": _normalize_text(raw_row.get("estado"), lower=False),
        "observaciones": _normalize_text(raw_row.get("observaciones"), lower=False),
    }

    if not normalized["id_lote"]:
        errors.append("id_lote es obligatorio")

    empresa = None
    unidad_operativa = None
    empresa_id_archivo = normalized["empresa_id"]

    if (
        empresa_activa is not None
        and empresa_id_archivo
        and empresa_id_archivo != empresa_activa.empresa_id
    ):
        warnings.append(TENANT_MISMATCH_WARNING)

    if normalized["unidad_id"]:
        unidad_operativa = _find_unit_in_empresa_scope(normalized["unidad_id"], empresa_activa)

        if unidad_operativa is None:
            if empresa_activa is not None and _find_unit_in_empresa_scope(normalized["unidad_id"]) is not None:
                errors.append("unidad_id no pertenece a la empresa activa")
            else:
                errors.append("unidad_id no existe")
        else:
            empresa = empresa_activa or unidad_operativa.empresa
            normalized["empresa_id"] = empresa.empresa_id
            normalized["empresa_aserradero"] = normalized["empresa_aserradero"] or empresa.nombre
            normalized["unidad_id"] = unidad_operativa.unidad_id
    elif empresa_activa is not None:
        empresa = empresa_activa
        normalized["empresa_id"] = empresa_activa.empresa_id
        normalized["empresa_aserradero"] = normalized["empresa_aserradero"] or empresa_activa.nombre
    elif normalized["empresa_id"]:
        empresa = Empresa.objects.filter(empresa_id=normalized["empresa_id"]).first()
        if empresa is None:
            errors.append("empresa_id no existe")
        else:
            normalized["empresa_aserradero"] = normalized["empresa_aserradero"] or empresa.nombre

    if empresa_activa is not None:
        empresa = empresa_activa
        normalized["empresa_id"] = empresa_activa.empresa_id
        normalized["empresa_aserradero"] = normalized["empresa_aserradero"] or empresa_activa.nombre

    normalized["empresa_obj"] = empresa
    normalized["unidad_operativa_obj"] = unidad_operativa

    if not normalized["empresa_aserradero"]:
        errors.append("empresa es obligatoria")
    if not normalized["especie"]:
        errors.append("especie es obligatoria")
    if not normalized["origen"]:
        errors.append("origen es obligatorio")

    try:
        normalized["fecha"] = _parse_lote_date(raw_row.get("fecha"))
    except ValueError as exc:
        errors.append(str(exc))

    try:
        volumen = _parse_lote_decimal(raw_row.get("volumen_m3"), "volumen_m3")
        if volumen is None:
            errors.append("volumen_m3 es obligatorio")
        else:
            normalized["volumen_m3"] = volumen
    except ValueError as exc:
        errors.append(str(exc))

    try:
        normalized["densidad_kg_m3"] = _parse_lote_decimal(
            raw_row.get("densidad_kg_m3"),
            "densidad_kg_m3",
        )
    except ValueError as exc:
        errors.append(str(exc))

    try:
        normalized["porcentaje_carbono"] = _parse_lote_decimal(
            raw_row.get("porcentaje_carbono"),
            "porcentaje_carbono",
        )
    except ValueError as exc:
        errors.append(str(exc))

    has_species_defaults = (
        EspecieMadera.objects.filter(nombre__iexact=normalized["especie"]).exists()
        if normalized["especie"]
        else False
    )
    has_carbon_overrides = (
        normalized.get("densidad_kg_m3") is not None
        and normalized.get("porcentaje_carbono") is not None
    )
    if normalized["especie"] and not has_species_defaults and not has_carbon_overrides:
        warnings.append("especie sin densidad conocida")

    return normalized, errors, warnings


def _parse_bool(value) -> bool:
    text = _normalize_text(value, lower=True)
    if text in {"0", "no", "false", "inactivo", "inactiva"}:
        return False
    return True


def _build_company_payload(raw_row: dict) -> tuple[dict, list[str]]:
    errors = []
    normalized = {
        "empresa_id": _normalize_text(raw_row.get("empresa_id"), lower=False).upper(),
        "nombre": _normalize_text(raw_row.get("nombre"), lower=False),
        "rut": _normalize_text(raw_row.get("rut"), lower=False),
        "region": _normalize_text(raw_row.get("region"), lower=False),
        "comuna": _normalize_text(raw_row.get("comuna"), lower=False),
        "direccion": _normalize_text(raw_row.get("direccion"), lower=False),
        "rubro": _normalize_text(raw_row.get("rubro"), lower=False),
        "email": _normalize_text(raw_row.get("email"), lower=False),
        "telefono": _normalize_text(raw_row.get("telefono"), lower=False),
        "contacto": _normalize_text(raw_row.get("contacto"), lower=False),
        "observaciones": _normalize_text(raw_row.get("observaciones"), lower=False),
    }

    if not normalized["empresa_id"]:
        errors.append("empresa_id es obligatorio")
    if not normalized["nombre"]:
        errors.append("nombre es obligatorio")

    return normalized, errors


def _build_unit_payload(raw_row: dict, empresa_activa=None) -> tuple[dict, list[str], list[str]]:
    errors = []
    warnings = []
    normalized = {
        "unidad_id": _normalize_text(raw_row.get("unidad_id"), lower=False).upper(),
        "empresa_id": _normalize_text(raw_row.get("empresa_id"), lower=False).upper(),
        "nombre": _normalize_text(raw_row.get("nombre"), lower=False),
        "tipo": _normalize_text(raw_row.get("tipo"), lower=False),
        "region": _normalize_text(raw_row.get("region"), lower=False),
        "comuna": _normalize_text(raw_row.get("comuna"), lower=False),
        "direccion": _normalize_text(raw_row.get("direccion"), lower=False),
        "descripcion": _normalize_text(raw_row.get("descripcion"), lower=False),
        "activa": _parse_bool(raw_row.get("activa")),
    }

    if not normalized["unidad_id"]:
        errors.append("unidad_id es obligatorio")
    if not normalized["nombre"]:
        errors.append("nombre es obligatorio")
    if not normalized["tipo"]:
        errors.append("tipo es obligatorio")

    empresa = None
    if empresa_activa is not None:
        if normalized["empresa_id"] and normalized["empresa_id"] != empresa_activa.empresa_id:
            warnings.append(TENANT_MISMATCH_WARNING)
        empresa = empresa_activa
        normalized["empresa_id"] = empresa_activa.empresa_id
    elif normalized["empresa_id"]:
        empresa = Empresa.objects.filter(empresa_id=normalized["empresa_id"]).first()
        if empresa is None:
            errors.append("empresa_id no existe")
    else:
        errors.append("empresa_id es obligatorio")

    normalized["empresa_obj"] = empresa

    allowed_types = {choice[0] for choice in UnidadOperativa.Tipo.choices}
    if normalized["tipo"] not in allowed_types:
        normalized["tipo"] = UnidadOperativa.Tipo.OTRO

    return normalized, errors, warnings


def _find_semantic_unit_duplicate(data: dict):
    empresa = data.get("empresa_obj")
    if not empresa or not data.get("nombre"):
        return None

    return (
        UnidadOperativa.objects.filter(
            empresa=empresa,
            nombre__iexact=data.get("nombre", ""),
            tipo=data.get("tipo", ""),
            region__iexact=data.get("region", ""),
            comuna__iexact=data.get("comuna", ""),
        )
        .exclude(unidad_id=data.get("unidad_id", ""))
        .first()
    )


def _lote_response_row(row: ParsedLoteRow) -> dict:
    data = {}
    for key, value in row.normalized.items():
        if key.endswith("_obj"):
            continue
        data[key] = str(value) if isinstance(value, Decimal) else value

    data["empresa"] = data.get("empresa_aserradero", "")

    return {
        "row_number": row.row_number,
        "status": row.status,
        "errors": row.errors,
        "warnings": row.warnings,
        "is_duplicate": row.is_duplicate,
        "exists_in_db": row.exists_in_db,
        "db_action": row.db_action,
        "raw": _stringify_payload(row.raw),
        "data": data,
    }


def _find_factor_for_activity(actividad: str, unidad: str):
    actividad_key = normalize_activity_key(actividad)
    return (
        FactorEmision.objects.filter(
            actividad_key=actividad_key,
            unidad__iexact=unidad,
        )
        .order_by("-anio", "fuente")
        .first()
    )


def _build_activity_payload(raw_row: dict, empresa_activa=None) -> tuple[dict, list[str], list[str]]:
    errors = []
    warnings = []
    id_lote = _normalize_text(raw_row.get("id_lote"), lower=False).upper()
    empresa_id = _normalize_text(raw_row.get("empresa_id"), lower=False).upper()
    empresa_id_archivo = empresa_id
    unidad_id = _normalize_text(raw_row.get("unidad_id"), lower=False).upper()
    actividad = format_activity_display_name(_normalize_text(raw_row.get("actividad"), lower=False))
    unidad = _normalize_text(raw_row.get("unidad"), lower=False)

    if not actividad:
        errors.append("actividad es requerida")
    if not unidad:
        errors.append("unidad es requerida")

    try:
        cantidad = _parse_lote_decimal(raw_row.get("cantidad"), "cantidad")
        if cantidad is None:
            errors.append("cantidad es requerida")
    except ValueError as exc:
        cantidad = None
        errors.append(str(exc))

    try:
        fecha = _parse_lote_date(raw_row.get("fecha"))
    except ValueError as exc:
        fecha = ""
        errors.append(str(exc))

    lote = None
    empresa = None
    unidad_operativa = None
    tipo_asignacion = EmisionLote.TipoAsignacion.EMPRESA

    if (
        empresa_activa is not None
        and empresa_id_archivo
        and empresa_id_archivo != empresa_activa.empresa_id
    ):
        warnings.append(TENANT_MISMATCH_WARNING)

    if id_lote:
        lote_global = (
            Lote.objects.select_related("empresa", "unidad_operativa", "unidad_operativa__empresa")
            .filter(id_lote=id_lote)
            .first()
        )
        if empresa_activa is not None:
            lote = _find_lote_in_empresa_scope(id_lote, empresa_activa)
        else:
            lote = lote_global

        if lote is None:
            if empresa_activa is not None and lote_global is not None:
                errors.append("id_lote no pertenece a la empresa activa")
            else:
                errors.append("id_lote no existe")
        else:
            empresa = empresa_activa or lote.empresa
            unidad_operativa = lote.unidad_operativa
            empresa_id = (
                (empresa_activa.empresa_id if empresa_activa is not None else empresa.empresa_id)
                if empresa
                else empresa_id
            )
            unidad_id = unidad_operativa.unidad_id if unidad_operativa else unidad_id
            tipo_asignacion = EmisionLote.TipoAsignacion.LOTE
    elif unidad_id:
        unidad_operativa = _find_unit_in_empresa_scope(unidad_id, empresa_activa)

        if unidad_operativa is None:
            if empresa_activa is not None and _find_unit_in_empresa_scope(unidad_id) is not None:
                errors.append("unidad_id no pertenece a la empresa activa")
            else:
                errors.append("unidad_id no existe")
        else:
            empresa = empresa_activa or unidad_operativa.empresa
            empresa_id = empresa.empresa_id
            unidad_id = unidad_operativa.unidad_id
            tipo_asignacion = EmisionLote.TipoAsignacion.UNIDAD
    elif empresa_activa is not None:
        empresa = empresa_activa
        empresa_id = empresa_activa.empresa_id
        tipo_asignacion = EmisionLote.TipoAsignacion.EMPRESA
    elif empresa_id:
        empresa = Empresa.objects.filter(empresa_id=empresa_id).first()
        if empresa is None:
            errors.append("empresa_id no existe")
        tipo_asignacion = EmisionLote.TipoAsignacion.EMPRESA
    else:
        errors.append("id_lote, unidad_id o empresa_id es requerido")

    if empresa_activa is not None:
        empresa = empresa_activa
        empresa_id = empresa_activa.empresa_id

    factor = _find_factor_for_activity(actividad, unidad) if actividad and unidad else None

    if actividad and unidad and factor is None:
        errors.append("factor de emision no encontrado")

    data = {
        "empresa_id": empresa_id,
        "unidad_id": unidad_id,
        "id_lote": id_lote,
        "actividad": actividad,
        "cantidad": cantidad,
        "unidad": unidad,
        "fecha": fecha,
        "factor_emision": factor.factor_emision if factor else None,
        "categoria": factor.categoria if factor else "",
        "fuente": factor.fuente if factor else "",
        "anio": factor.anio if factor else "",
        "tipo_asignacion": tipo_asignacion,
    }

    return data, errors, warnings


def _activity_key(data: dict) -> tuple:
    return (
        data.get("empresa_id"),
        data.get("unidad_id"),
        data.get("id_lote"),
        _normalize_text(data.get("actividad"), lower=True),
        _normalize_text(data.get("unidad"), lower=True),
        data.get("fecha"),
        str(data.get("cantidad")),
        str(data.get("factor_emision")),
    )


def _activity_exists(data: dict) -> bool:
    lote = Lote.objects.filter(id_lote=data.get("id_lote")).first()
    empresa = Empresa.objects.filter(empresa_id=data.get("empresa_id")).first()
    unidad_operativa = UnidadOperativa.objects.filter(unidad_id=data.get("unidad_id")).first()
    filters = dict(
        actividad__iexact=data.get("actividad"),
        unidad__iexact=data.get("unidad"),
        fecha=data.get("fecha"),
        cantidad=data.get("cantidad"),
        factor_emision=data.get("factor_emision"),
    )
    if lote:
        filters["lote"] = lote
    elif unidad_operativa:
        filters["unidad_operativa"] = unidad_operativa
        filters["lote__isnull"] = True
    elif empresa:
        filters["empresa"] = empresa
        filters["lote__isnull"] = True
        filters["unidad_operativa__isnull"] = True
    else:
        return False

    return EmisionLote.objects.filter(**filters).exists()


def _activity_response_data(data: dict) -> dict:
    return {
        **data,
        "cantidad": str(data.get("cantidad", "")) if data.get("cantidad") is not None else "",
        "factor_emision": (
            str(data.get("factor_emision", ""))
            if data.get("factor_emision") is not None
            else ""
        ),
    }


class ImportadorFactores:
    @staticmethod
    def previsualizar(uploaded_file) -> dict:
        raw_rows, file_format = read_uploaded_factor_rows(uploaded_file)
        batch_id = uuid.uuid4().hex

        seen_keys: dict[tuple, int] = {}
        parsed_rows: list[ParsedFactorRow] = []

        for raw_row in raw_rows:
            normalized, errors = _build_row_payload(raw_row)
            key = _row_lookup_key(normalized)
            is_duplicate = False
            exists_in_db = False
            db_action = None

            if not errors and all(key):
                if key in seen_keys:
                    is_duplicate = True
                    errors.append(f"Fila duplicada de la fila {seen_keys[key]}")
                else:
                    seen_keys[key] = raw_row["row_number"]

                exists_in_db = FactorEmision.objects.filter(
                    actividad_key=normalized["actividad_key"],
                    unidad__iexact=normalized["unidad"],
                    factor_emision=normalized["factor_emision"],
                    anio=normalized["anio"],
                ).exists() or FactorEmision.objects.filter(
                    actividad_key=normalized["actividad_key"],
                    unidad__iexact=normalized["unidad"],
                    fuente__iexact=normalized["fuente"],
                    anio=normalized["anio"],
                ).exists()
                db_action = "actualizar" if exists_in_db else "crear"

            status = "valid" if not errors and not is_duplicate else "error"
            parsed_rows.append(
                ParsedFactorRow(
                    row_number=raw_row["row_number"],
                    raw={k: v for k, v in raw_row.items() if k != "row_number"},
                    normalized=normalized,
                    errors=errors,
                    status=status,
                    is_duplicate=is_duplicate,
                    exists_in_db=exists_in_db,
                    db_action=db_action,
                )
            )

        result_rows = [
            {
                "row_number": row.row_number,
                "status": row.status,
                "errors": row.errors,
                "warnings": row.normalized.get("observaciones", []),
                "is_duplicate": row.is_duplicate,
                "exists_in_db": row.exists_in_db,
                "db_action": row.db_action,
                "data": {
                    **row.normalized,
                    "factor_emision": str(row.normalized.get("factor_emision", "")),
                },
            }
            for row in parsed_rows
        ]

        summary = {
            "total_filas": len(result_rows),
            "validas": sum(1 for row in parsed_rows if row.status == "valid"),
            "con_error": sum(1 for row in parsed_rows if row.status == "error"),
            "duplicadas": sum(1 for row in parsed_rows if row.is_duplicate),
            "posibles_actualizaciones": sum(
                1 for row in parsed_rows if row.status == "valid" and row.exists_in_db
            ),
            "posibles_creaciones": sum(
                1 for row in parsed_rows if row.status == "valid" and not row.exists_in_db
            ),
        }

        cache.set(
            f"{FACTOR_IMPORT_CACHE_PREFIX}{batch_id}",
            {
                "format": file_format,
                "rows": result_rows,
                "summary": summary,
            },
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )

        logger.info(
            "Preview import factores batch_id=%s filas=%s validas=%s errores=%s",
            batch_id,
            summary["total_filas"],
            summary["validas"],
            summary["con_error"],
        )

        return {"batch_id": batch_id, "summary": summary, "rows": result_rows}

    @staticmethod
    def confirmar(rows: list[dict] | None = None, batch_id: str | None = None) -> dict:
        if batch_id:
            cached = cache.get(f"{FACTOR_IMPORT_CACHE_PREFIX}{batch_id}")
            if not cached:
                raise ValueError("El batch_id no existe o expiró")
            rows = cached.get("rows", [])

        if not rows:
            raise ValueError("No hay filas validas para importar")

        created = 0
        updated = 0
        rejected = 0
        errors: list[dict] = []

        with transaction.atomic():
            for row in rows:
                data = row.get("data") or row
                row_number = row.get("row_number")

                try:
                    normalized, validation_errors = _build_row_payload(data)
                    if validation_errors:
                        rejected += 1
                        errors.append({"row_number": row_number, "errors": validation_errors})
                        continue

                    factor = FactorEmision.objects.filter(
                        actividad_key=normalized["actividad_key"],
                        unidad__iexact=normalized["unidad"],
                        fuente__iexact=normalized["fuente"],
                        anio=normalized["anio"],
                    ).first()
                    if factor is None:
                        factor = FactorEmision.objects.filter(
                            actividad_key=normalized["actividad_key"],
                            unidad__iexact=normalized["unidad"],
                            factor_emision=normalized["factor_emision"],
                            anio=normalized["anio"],
                        ).first()
                    was_created = factor is None

                    if factor is None:
                        factor = FactorEmision(
                            actividad_key=normalized["actividad_key"],
                            unidad=normalized["unidad"],
                            fuente=normalized["fuente"],
                            anio=normalized["anio"],
                        )

                    if was_created or _is_readable_activity_name(normalized["actividad"]):
                        factor.actividad = normalized["actividad"]
                        factor.fuente = normalized["fuente"]
                    factor.categoria = normalized["categoria"]
                    factor.descripcion = normalized["descripcion"]
                    factor.metadata_clasificacion = normalized.get(
                        "metadata_clasificacion",
                        {},
                    )
                    factor.factor_emision = normalized["factor_emision"]
                    try:
                        with transaction.atomic():
                            factor.save()
                    except IntegrityError:
                        rejected += 1
                        errors.append(
                            {
                                "row_number": row_number,
                                "errors": [
                                    "Ya existe un factor de emision con la misma actividad, unidad, fuente y anio."
                                ],
                            }
                        )
                        continue
                    created += 1 if was_created else 0
                    updated += 0 if was_created else 1
                except Exception:
                    logger.exception("Error al confirmar factor de emision")
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row_number,
                            "errors": [
                                "No se pudo guardar el factor de emision. Revisa sus datos y duplicados."
                            ],
                        }
                    )

        summary = {
            "creados": created,
            "actualizados": updated,
            "rechazados": rejected,
            "errores": errors,
        }
        logger.info(
            "Confirm import factores batch_id=%s creados=%s actualizados=%s rechazados=%s",
            batch_id,
            created,
            updated,
            rejected,
        )
        return summary


class ImportadorEmpresas:
    @staticmethod
    def previsualizar(uploaded_file) -> dict:
        raw_rows, file_format = read_uploaded_company_rows(uploaded_file)
        batch_id = uuid.uuid4().hex
        seen_ids: dict[str, int] = {}
        result_rows = []

        for raw_row in raw_rows:
            data, errors = _build_company_payload(raw_row)
            empresa_id = data.get("empresa_id")
            is_duplicate = False
            exists_in_db = False
            db_action = None

            if empresa_id:
                if empresa_id in seen_ids:
                    is_duplicate = True
                    errors.append(f"empresa_id duplicado de la fila {seen_ids[empresa_id]}")
                else:
                    seen_ids[empresa_id] = raw_row["row_number"]

                exists_in_db = Empresa.objects.filter(empresa_id=empresa_id).exists()
                db_action = "actualizar" if exists_in_db else "crear"

            status = "valid" if not errors and not is_duplicate else "error"
            result_rows.append(
                {
                    "row_number": raw_row["row_number"],
                    "status": status,
                    "errors": errors,
                    "is_duplicate": is_duplicate,
                    "exists_in_db": exists_in_db,
                    "db_action": db_action,
                    "data": data,
                }
            )

        summary = {
            "total_filas": len(result_rows),
            "validas": sum(1 for row in result_rows if row["status"] == "valid"),
            "con_error": sum(1 for row in result_rows if row["status"] == "error"),
            "duplicadas": sum(1 for row in result_rows if row["is_duplicate"]),
            "posibles_creaciones": sum(
                1 for row in result_rows if row["status"] == "valid" and not row["exists_in_db"]
            ),
            "posibles_actualizaciones": sum(
                1 for row in result_rows if row["status"] == "valid" and row["exists_in_db"]
            ),
        }
        cache.set(
            f"{COMPANY_IMPORT_CACHE_PREFIX}{batch_id}",
            {"format": file_format, "rows": result_rows, "summary": summary},
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )
        return {"batch_id": batch_id, "summary": summary, "rows": result_rows}

    @staticmethod
    def confirmar(rows: list[dict] | None = None, batch_id: str | None = None) -> dict:
        if batch_id:
            cached = cache.get(f"{COMPANY_IMPORT_CACHE_PREFIX}{batch_id}")
            if not cached:
                raise ValueError("El batch_id no existe o expiro")
            rows = cached.get("rows", [])

        if not rows:
            raise ValueError("No hay filas validas para importar")

        created = 0
        updated = 0
        rejected = 0
        errors = []

        with transaction.atomic():
            for row in rows:
                if row.get("status") and row.get("status") != "valid":
                    rejected += 1
                    errors.append({"row_number": row.get("row_number"), "errors": row.get("errors")})
                    continue

                data, validation_errors = _build_company_payload(row.get("data") or row)
                if validation_errors:
                    rejected += 1
                    errors.append({"row_number": row.get("row_number"), "errors": validation_errors})
                    continue

                _, was_created = Empresa.objects.update_or_create(
                    empresa_id=data["empresa_id"],
                    defaults={
                        "nombre": data["nombre"],
                        "rut": data["rut"],
                        "region": data["region"],
                        "comuna": data["comuna"],
                        "direccion": data["direccion"],
                        "rubro": data["rubro"],
                        "email": data["email"],
                        "telefono": data["telefono"],
                        "contacto": data["contacto"],
                        "observaciones": data["observaciones"],
                    },
                )
                created += 1 if was_created else 0
                updated += 0 if was_created else 1

        return {
            "creados": created,
            "actualizados": updated,
            "rechazados": rejected,
            "errores": errors,
        }


class ImportadorUnidadesOperativas:
    @staticmethod
    def previsualizar(uploaded_file, empresa_activa=None) -> dict:
        raw_rows, file_format = read_uploaded_unit_rows(uploaded_file, empresa_activa=empresa_activa)
        batch_id = uuid.uuid4().hex
        seen_ids: dict[str, int] = {}
        result_rows = []

        for raw_row in raw_rows:
            data, errors, warnings = _build_unit_payload(raw_row, empresa_activa=empresa_activa)
            unidad_id = data.get("unidad_id")
            is_duplicate = False
            exists_in_db = False
            db_action = None

            if unidad_id:
                if unidad_id in seen_ids:
                    is_duplicate = True
                    errors.append(f"unidad_id duplicado de la fila {seen_ids[unidad_id]}")
                else:
                    seen_ids[unidad_id] = raw_row["row_number"]

                unidad_existente = UnidadOperativa.objects.select_related("empresa").filter(
                    unidad_id=unidad_id
                ).first()
                if unidad_existente is not None:
                    if (
                        empresa_activa is not None
                        and unidad_existente.empresa_id != empresa_activa.id
                    ):
                        errors.append("unidad_id ya existe en otra empresa")
                    else:
                        exists_in_db = True
                        db_action = "actualizar"
                else:
                    db_action = "crear"

            semantic_duplicate = _find_semantic_unit_duplicate(data)
            if semantic_duplicate is not None:
                errors.append(
                    "Ya existe una unidad operativa con el mismo nombre, tipo, region y comuna "
                    f"en esta empresa ({semantic_duplicate.unidad_id})."
                )

            status = "valid" if not errors and not is_duplicate else "error"
            result_rows.append(
                {
                    "row_number": raw_row["row_number"],
                    "status": status,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "exists_in_db": exists_in_db,
                    "db_action": db_action,
                    "data": {
                        key: value
                        for key, value in data.items()
                        if not key.endswith("_obj")
                    },
                }
            )

        summary = {
            "total_filas": len(result_rows),
            "validas": sum(1 for row in result_rows if row["status"] == "valid"),
            "con_error": sum(1 for row in result_rows if row["status"] == "error"),
            "duplicadas": sum(1 for row in result_rows if row["is_duplicate"]),
            "posibles_creaciones": sum(
                1 for row in result_rows if row["status"] == "valid" and not row["exists_in_db"]
            ),
            "posibles_actualizaciones": sum(
                1 for row in result_rows if row["status"] == "valid" and row["exists_in_db"]
            ),
        }
        cache.set(
            f"{UNIT_IMPORT_CACHE_PREFIX}{batch_id}",
            {"format": file_format, "rows": result_rows, "summary": summary},
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )
        return {"batch_id": batch_id, "summary": summary, "rows": result_rows}

    @staticmethod
    def confirmar(rows: list[dict] | None = None, batch_id: str | None = None, empresa_activa=None) -> dict:
        if batch_id:
            cached = cache.get(f"{UNIT_IMPORT_CACHE_PREFIX}{batch_id}")
            if not cached:
                raise ValueError("El batch_id no existe o expiro")
            rows = cached.get("rows", [])

        if not rows:
            raise ValueError("No hay filas validas para importar")

        created = 0
        updated = 0
        rejected = 0
        errors = []

        with transaction.atomic():
            for row in rows:
                if row.get("status") and row.get("status") != "valid":
                    rejected += 1
                    errors.append({"row_number": row.get("row_number"), "errors": row.get("errors")})
                    continue

                data, validation_errors, _warnings = _build_unit_payload(
                    row.get("data") or row,
                    empresa_activa=empresa_activa,
                )
                if validation_errors:
                    rejected += 1
                    errors.append({"row_number": row.get("row_number"), "errors": validation_errors})
                    continue

                semantic_duplicate = _find_semantic_unit_duplicate(data)
                if semantic_duplicate is not None:
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row.get("row_number"),
                            "errors": [
                                "Ya existe una unidad operativa con el mismo nombre, tipo, region y comuna "
                                f"en esta empresa ({semantic_duplicate.unidad_id})."
                            ],
                        }
                    )
                    continue

                unidad_existente = UnidadOperativa.objects.select_related("empresa").filter(
                    unidad_id=data["unidad_id"]
                ).first()
                if (
                    unidad_existente is not None
                    and empresa_activa is not None
                    and unidad_existente.empresa_id != empresa_activa.id
                ):
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row.get("row_number"),
                            "errors": ["unidad_id ya existe en otra empresa"],
                        }
                    )
                    continue

                if unidad_existente is None:
                    UnidadOperativa.objects.create(
                        unidad_id=data["unidad_id"],
                        empresa=data["empresa_obj"],
                        nombre=data["nombre"],
                        tipo=data["tipo"],
                        region=data["region"],
                        comuna=data["comuna"],
                        direccion=data["direccion"],
                        descripcion=data["descripcion"],
                        activa=data["activa"],
                    )
                    was_created = True
                else:
                    unidad_existente.empresa = data["empresa_obj"]
                    unidad_existente.nombre = data["nombre"]
                    unidad_existente.tipo = data["tipo"]
                    unidad_existente.region = data["region"]
                    unidad_existente.comuna = data["comuna"]
                    unidad_existente.direccion = data["direccion"]
                    unidad_existente.descripcion = data["descripcion"]
                    unidad_existente.activa = data["activa"]
                    unidad_existente.save()
                    was_created = False

                created += 1 if was_created else 0
                updated += 0 if was_created else 1

        return {
            "creados": created,
            "actualizados": updated,
            "rechazados": rejected,
            "errores": errors,
        }


class ImportadorLotes:
    @staticmethod
    def previsualizar(uploaded_file, empresa_activa=None) -> dict:
        raw_rows, file_format = read_uploaded_lote_rows(uploaded_file)
        batch_id = uuid.uuid4().hex
        seen_ids: dict[str, int] = {}
        parsed_rows: list[ParsedLoteRow] = []

        for raw_row in raw_rows:
            normalized, errors, warnings = _build_lote_payload(raw_row, empresa_activa=empresa_activa)
            id_lote = normalized.get("id_lote")
            is_duplicate = False
            exists_in_db = False
            db_action = None

            if id_lote:
                if id_lote in seen_ids:
                    is_duplicate = True
                    errors.append(f"id_lote duplicado de la fila {seen_ids[id_lote]}")
                else:
                    seen_ids[id_lote] = raw_row["row_number"]

                lote_existente = (
                    Lote.objects.select_related("empresa", "unidad_operativa", "unidad_operativa__empresa")
                    .filter(id_lote=id_lote)
                    .first()
                )
                if lote_existente is not None:
                    if (
                        empresa_activa is not None
                        and not _lote_belongs_to_empresa(lote_existente, empresa_activa)
                    ):
                        errors.append("id_lote ya existe en otra empresa")
                    else:
                        exists_in_db = True
                        db_action = "actualizar"
                else:
                    db_action = "crear"

            status = "valid" if not errors and not is_duplicate else "error"
            parsed_rows.append(
                ParsedLoteRow(
                    row_number=raw_row["row_number"],
                    raw={k: v for k, v in raw_row.items() if k != "row_number"},
                    normalized=normalized,
                    errors=errors,
                    warnings=warnings,
                    status=status,
                    is_duplicate=is_duplicate,
                    exists_in_db=exists_in_db,
                    db_action=db_action,
                )
            )

        result_rows = [_lote_response_row(row) for row in parsed_rows]
        summary = {
            "total_filas": len(result_rows),
            "validas": sum(1 for row in parsed_rows if row.status == "valid"),
            "con_error": sum(1 for row in parsed_rows if row.status == "error"),
            "duplicadas": sum(1 for row in parsed_rows if row.is_duplicate),
            "posibles_creaciones": sum(
                1 for row in parsed_rows if row.status == "valid" and not row.exists_in_db
            ),
            "posibles_actualizaciones": sum(
                1 for row in parsed_rows if row.status == "valid" and row.exists_in_db
            ),
            "lotes_nuevos": sum(
                1 for row in parsed_rows if row.status == "valid" and not row.exists_in_db
            ),
            "lotes_existentes": sum(
                1 for row in parsed_rows if row.status == "valid" and row.exists_in_db
            ),
        }

        cache.set(
            f"{LOTE_IMPORT_CACHE_PREFIX}{batch_id}",
            {
                "format": file_format,
                "rows": result_rows,
                "summary": summary,
            },
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )

        logger.info(
            "Preview import lotes batch_id=%s filas=%s validas=%s errores=%s",
            batch_id,
            summary["total_filas"],
            summary["validas"],
            summary["con_error"],
        )

        return {"batch_id": batch_id, "summary": summary, "rows": result_rows}

    @staticmethod
    def confirmar(rows: list[dict] | None = None, batch_id: str | None = None, empresa_activa=None) -> dict:
        if batch_id:
            cached = cache.get(f"{LOTE_IMPORT_CACHE_PREFIX}{batch_id}")
            if not cached:
                raise ValueError("El batch_id no existe o expiro")
            rows = cached.get("rows", [])

        if not rows:
            raise ValueError("No hay filas validas para importar")

        created = 0
        updated = 0
        rejected = 0
        errors: list[dict] = []

        with transaction.atomic():
            for row in rows:
                if row.get("status") and row.get("status") != "valid":
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row.get("row_number"),
                            "errors": row.get("errors") or ["fila no valida"],
                        }
                    )
                    continue

                data = row.get("data") or row
                row_number = row.get("row_number")

                try:
                    raw = row.get("raw") or data
                    normalized, validation_errors, warnings = _build_lote_payload(data, empresa_activa=empresa_activa)
                    if validation_errors:
                        rejected += 1
                        errors.append({"row_number": row_number, "errors": validation_errors})
                        continue

                    lote_existente = (
                        Lote.objects.select_related("empresa", "unidad_operativa", "unidad_operativa__empresa")
                        .filter(id_lote=normalized["id_lote"])
                        .first()
                    )
                    if (
                        lote_existente is not None
                        and empresa_activa is not None
                        and not _lote_belongs_to_empresa(lote_existente, empresa_activa)
                    ):
                        rejected += 1
                        errors.append({"row_number": row_number, "errors": ["id_lote ya existe en otra empresa"]})
                        continue

                    defaults = {
                        "empresa": normalized.get("empresa_obj"),
                        "unidad_operativa": normalized.get("unidad_operativa_obj"),
                        "empresa_aserradero": normalized["empresa_aserradero"],
                        "fecha": normalized["fecha"],
                        "especie": normalized["especie"],
                        "volumen_m3": normalized["volumen_m3"],
                        "origen": normalized["origen"],
                        "tipo_producto": normalized.get("tipo_producto") or "",
                        "densidad_kg_m3": normalized.get("densidad_kg_m3"),
                        "porcentaje_carbono": normalized.get("porcentaje_carbono"),
                        "estado": normalized.get("estado") or "",
                        "observaciones": normalized.get("observaciones") or "",
                    }
                    if lote_existente is None:
                        lote = Lote.objects.create(id_lote=normalized["id_lote"], **defaults)
                        was_created = True
                    else:
                        for field, value in defaults.items():
                            setattr(lote_existente, field, value)
                        lote_existente.save()
                        lote = lote_existente
                        was_created = False

                    created += 1 if was_created else 0
                    updated += 0 if was_created else 1

                    HistorialCambioLote.objects.create(
                        lote=lote,
                        tipo=HistorialCambioLote.TipoCambio.IMPORTADO,
                        fuente="importador_lotes",
                        raw_payload=_stringify_payload(raw),
                        normalized_payload=_stringify_payload(normalized),
                        metadata={
                            "accion": "creado" if was_created else "actualizado",
                            "warnings": warnings,
                        },
                    )
                except Exception:
                    logger.exception("Error al confirmar lote")
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row_number,
                            "errors": [
                                "No se pudo guardar el lote. Revisa sus datos y referencias."
                            ],
                        }
                    )

        summary = {
            "creados": created,
            "actualizados": updated,
            "rechazados": rejected,
            "errores": errors,
        }
        logger.info(
            "Confirm import lotes batch_id=%s creados=%s actualizados=%s rechazados=%s",
            batch_id,
            created,
            updated,
            rejected,
        )
        return summary


class ImportadorActividadesLote:
    @staticmethod
    def previsualizar(uploaded_file, empresa_activa=None) -> dict:
        raw_rows, file_format = read_uploaded_activity_rows(uploaded_file)
        batch_id = uuid.uuid4().hex
        seen_keys: dict[tuple, int] = {}
        result_rows = []

        for raw_row in raw_rows:
            data, errors, warnings = _build_activity_payload(raw_row, empresa_activa=empresa_activa)
            key = _activity_key(data)
            is_duplicate = False
            exists_in_db = False
            db_action = None

            if not errors and all(key):
                if key in seen_keys:
                    is_duplicate = True
                    errors.append(f"Fila duplicada de la fila {seen_keys[key]}")
                else:
                    seen_keys[key] = raw_row["row_number"]

                exists_in_db = _activity_exists(data)
                if exists_in_db:
                    is_duplicate = True
                    db_action = "omitir"
                else:
                    db_action = "crear"

            status = "valid" if not errors and not is_duplicate else "duplicate" if is_duplicate else "error"
            result_rows.append(
                {
                    "row_number": raw_row["row_number"],
                    "status": status,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "exists_in_db": exists_in_db,
                    "db_action": db_action,
                    "factor_found": bool(data.get("factor_emision")),
                    "lote_exists": (
                        (
                            _find_lote_in_empresa_scope(data.get("id_lote"), empresa_activa)
                            is not None
                            if empresa_activa is not None and data.get("id_lote")
                            else Lote.objects.filter(id_lote=data.get("id_lote")).exists()
                        )
                        if data.get("id_lote")
                        else False
                    ),
                    "data": _activity_response_data(data),
                }
            )

        summary = {
            "filas_validas": sum(1 for row in result_rows if row["status"] == "valid"),
            "filas_con_error": sum(1 for row in result_rows if row["status"] == "error"),
            "duplicados": sum(1 for row in result_rows if row["is_duplicate"]),
            "posibles_creaciones": sum(
                1 for row in result_rows if row["status"] == "valid" and row["db_action"] == "crear"
            ),
            "posibles_actualizaciones": 0,
            "factores_encontrados": sum(1 for row in result_rows if row["factor_found"]),
            "factores_faltantes": sum(1 for row in result_rows if not row["factor_found"]),
            "lotes_encontrados": sum(1 for row in result_rows if row["lote_exists"]),
            "lotes_nuevos_detectados": sum(
                1
                for row in result_rows
                if row["data"].get("id_lote") and not row["lote_exists"]
            ),
        }
        cache.set(
            f"{ACTIVITY_IMPORT_CACHE_PREFIX}{batch_id}",
            {
                "format": file_format,
                "rows": result_rows,
                "summary": summary,
                "confirmed": False,
            },
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )

        return {"batch_id": batch_id, "summary": summary, "rows": result_rows}

    @staticmethod
    def confirmar(rows: list[dict] | None = None, batch_id: str | None = None, empresa_activa=None) -> dict:
        cache_key = f"{ACTIVITY_IMPORT_CACHE_PREFIX}{batch_id}" if batch_id else None

        if batch_id:
            cached = cache.get(cache_key)
            if not cached:
                raise ValueError("El batch_id no existe o expiro")
            if cached.get("confirmed"):
                return {
                    "creados": 0,
                    "created": 0,
                    "actualizados": 0,
                    "omitidos": 0,
                    "duplicados": 0,
                    "rechazados": 0,
                    "errores": [],
                    "message": "Esta importación ya fue aplicada.",
                }
            rows = cached.get("rows", [])

        if not rows:
            raise ValueError("No hay filas validas para importar")

        created = 0
        updated = 0
        omitted = 0
        duplicated = 0
        rejected = 0
        errors: list[dict] = []

        with transaction.atomic():
            for row in rows:
                row_number = row.get("row_number")

                if row.get("status") in {"duplicate"}:
                    duplicated += 1
                    omitted += 1
                    continue

                if row.get("status") and row.get("status") != "valid":
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row_number,
                            "errors": row.get("errors") or ["fila no valida"],
                        }
                    )
                    continue

                data, validation_errors, _warnings = _build_activity_payload(
                    row.get("data") or row,
                    empresa_activa=empresa_activa,
                )
                if validation_errors:
                    rejected += 1
                    errors.append({"row_number": row_number, "errors": validation_errors})
                    continue

                if _activity_exists(data):
                    duplicated += 1
                    omitted += 1
                    continue

                try:
                    with transaction.atomic():
                        lote = (
                            Lote.objects.get(id_lote=data["id_lote"])
                            if data.get("id_lote")
                            else None
                        )
                        empresa = (
                            Empresa.objects.get(empresa_id=data["empresa_id"])
                            if data.get("empresa_id")
                            else None
                        )
                        unidad_operativa = (
                            UnidadOperativa.objects.get(unidad_id=data["unidad_id"])
                            if data.get("unidad_id")
                            else None
                        )
                        EmisionLote.objects.create(
                            lote=lote,
                            empresa=empresa,
                            unidad_operativa=unidad_operativa,
                            actividad=data["actividad"],
                            categoria=data.get("categoria") or "",
                            cantidad=data["cantidad"],
                            unidad=data["unidad"],
                            fecha=data["fecha"],
                            factor_emision=data["factor_emision"],
                        )
                    created += 1
                except IntegrityError:
                    duplicated += 1
                    omitted += 1
                except Exception:
                    logger.exception("Error al confirmar actividad")
                    rejected += 1
                    errors.append(
                        {
                            "row_number": row_number,
                            "errors": [
                                "No se pudo guardar la actividad. Revisa sus datos y referencias."
                            ],
                        }
                    )

        if cache_key:
            cached = cache.get(cache_key) or {}
            cached["confirmed"] = True
            cache.set(cache_key, cached, timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS)

        return {
            "creados": created,
            "created": created,
            "actualizados": updated,
            "omitidos": omitted,
            "duplicados": duplicated,
            "rechazados": rejected,
            "errores": errors,
        }

    @staticmethod
    def importar_para_lote(lote: Lote, rows: Iterable[dict]):
        created = 0
        omitted = 0
        for r in rows:
            data = {
                "id_lote": lote.id_lote,
                "actividad": r.get("actividad"),
                "cantidad": r.get("cantidad") or 0,
                "unidad": r.get("unidad") or "",
                "fecha": r.get("fecha"),
                "factor_emision": r.get("factor_emision") or 0,
            }

            if _activity_exists(data):
                omitted += 1
                continue

            EmisionLote.objects.create(
                lote=lote,
                actividad=data["actividad"],
                cantidad=data["cantidad"],
                unidad=data["unidad"],
                fecha=data["fecha"],
                factor_emision=data["factor_emision"],
            )
            created += 1
        return {"importadas": created, "omitidas": omitted}


__all__ = [
    "ImportadorEmpresas",
    "ImportadorFactores",
    "ImportadorLotes",
    "ImportadorActividadesLote",
    "read_uploaded_company_rows",
    "read_uploaded_factor_rows",
    "read_uploaded_lote_rows",
    "read_uploaded_activity_rows",
]
