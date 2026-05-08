"""Importador de empresa completa desde XLSX con múltiples hojas."""

import logging
import uuid
from decimal import Decimal
from pathlib import Path

from django.core.cache import cache
from django.db import IntegrityError
from openpyxl import load_workbook

from ..factores import format_activity_display_name, normalize_activity_key
from ..models import Empresa, Lote, EmisionLote, normalize_identifier
from .importadores import (
    _build_company_payload,
    _build_unit_payload,
    _build_lote_payload,
    _build_activity_payload,
    _build_row_payload,
    _parse_lote_date,
    _parse_lote_decimal,
    _normalize_header_key,
    FACTOR_IMPORT_CACHE_TTL_SECONDS,
    _normalize_text,
    ImportadorEmpresas,
    ImportadorActividadesLote,
    ImportadorFactores,
    ImportadorLotes,
    ImportadorUnidadesOperativas,
)

logger = logging.getLogger(__name__)

COMPLETE_IMPORT_CACHE_PREFIX = "empresa_completa_import_batch:"


def _read_xlsx_sheet(uploaded_file, sheet_name: str) -> list[dict]:
    """Lee una hoja específica de un archivo XLSX."""
    try:
        wb = load_workbook(uploaded_file)
        ws = wb[sheet_name]
        rows = []
        headers = None
        row_number = 0

        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:
                headers = [h for h in row if h]
                continue
            if not headers or not any(row):
                continue

            row_number += 1
            data = {"row_number": row_number}
            for header, cell in zip(headers, row):
                if header:
                    data[_normalize_text(header, lower=True)] = cell

            rows.append(data)

        return rows
    except KeyError:
        return []
    except Exception:
        raise ValueError(f"No se pudo leer la hoja {sheet_name}. Revisa el formato del archivo.")


def _read_normalized_xlsx_sheet(uploaded_file, sheet_name: str) -> list[dict]:
    try:
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
        try:
            ws = wb[sheet_name]
            rows = []
            headers = None
            for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if idx == 1:
                    headers = [_normalize_header_key(h) if h is not None else "" for h in row]
                    continue
                if not headers or not any(row):
                    continue

                values = list(row or [])
                data = {"row_number": idx}
                for index, header in enumerate(headers):
                    if header:
                        data[header] = values[index] if index < len(values) else None
                rows.append(data)
            return rows
        finally:
            wb.close()
    except KeyError:
        return []
    except Exception:
        raise ValueError(f"No se pudo leer la hoja {sheet_name}. Revisa el formato del archivo.")


def _read_normalized_xlsx_sheets(uploaded_file, sheet_names: list[str]) -> dict[str, list[dict]]:
    try:
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
        try:
            result = {}
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    result[sheet_name] = []
                    continue

                ws = wb[sheet_name]
                rows = []
                headers = None
                for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if idx == 1:
                        headers = [_normalize_header_key(h) if h is not None else "" for h in row]
                        continue
                    if not headers or not any(row):
                        continue

                    values = list(row or [])
                    data = {"row_number": idx}
                    for index, header in enumerate(headers):
                        if header:
                            data[header] = values[index] if index < len(values) else None
                    rows.append(data)
                result[sheet_name] = rows
            return result
        finally:
            wb.close()
    except Exception as exc:
        logger.exception("[IMPORT_COMPLETE] No se pudo leer el archivo XLSX")
        raise ValueError("No se pudo leer el archivo XLSX. Revisa que no este danado y que tenga las hojas requeridas.") from exc


def _strip_runtime_objects(data: dict) -> dict:
    return {key: value for key, value in (data or {}).items() if not key.endswith("_obj")}


def _remove_error(errors: list[str], message: str) -> None:
    while message in errors:
        errors.remove(message)


def _factor_lookup_key(actividad: str, unidad: str) -> tuple[str, str]:
    return (normalize_activity_key(actividad), _normalize_text(unidad, lower=True))


def _build_factor_lookup(rows: list[dict]) -> dict[tuple[str, str], dict]:
    lookup = {}
    for row in rows:
        data = row.get("data") or {}
        if row.get("status") != "valid":
            continue
        key = _factor_lookup_key(data.get("actividad"), data.get("unidad"))
        if all(key):
            lookup[key] = data
    return lookup


def _apply_file_factor(data: dict, errors: list[str], factor_lookup: dict[tuple[str, str], dict]) -> None:
    if data.get("factor_emision"):
        return

    factor = factor_lookup.get(_factor_lookup_key(data.get("actividad"), data.get("unidad")))
    if not factor:
        return

    data["factor_emision"] = Decimal(str(factor.get("factor_emision")))
    data["categoria"] = factor.get("categoria", "")
    data["fuente"] = factor.get("fuente", "")
    data["anio"] = factor.get("anio", "")
    _remove_error(errors, "factor de emision no encontrado")


def _build_complete_activity_payload(raw_row: dict, empresa_activa=None) -> tuple[dict, list[str], list[str]]:
    errors = []
    warnings = []
    id_lote = _normalize_text(raw_row.get("id_lote"), lower=False).upper()
    empresa_id = _normalize_text(raw_row.get("empresa_id"), lower=False).upper()
    unidad_id = _normalize_text(raw_row.get("unidad_id"), lower=False).upper()
    actividad = format_activity_display_name(_normalize_text(raw_row.get("actividad"), lower=False))
    unidad = _normalize_text(raw_row.get("unidad"), lower=False)

    if not actividad:
        errors.append("actividad es requerida")
    if not unidad:
        errors.append("unidad es requerida")

    try:
        cantidad = _parse_lote_decimal(raw_row.get("cantidad"), "cantidad")
        if cantidad is None:
            errors.append("cantidad es requerida")
    except ValueError as exc:
        cantidad = None
        errors.append(str(exc))

    try:
        fecha = _parse_lote_date(raw_row.get("fecha"))
    except ValueError as exc:
        fecha = ""
        errors.append(str(exc))

    if empresa_activa is not None:
        empresa_id = empresa_activa.empresa_id

    return {
        "empresa_id": empresa_id,
        "unidad_id": unidad_id,
        "id_lote": id_lote,
        "actividad": actividad,
        "cantidad": cantidad,
        "unidad": unidad,
        "fecha": fecha,
        "factor_emision": None,
        "categoria": "",
        "fuente": "",
        "anio": "",
        "tipo_asignacion": EmisionLote.TipoAsignacion.LOTE if id_lote else EmisionLote.TipoAsignacion.EMPRESA,
    }, errors, warnings


def _confirm_section(section_name: str, confirm_fn, **kwargs) -> dict:
    try:
        return confirm_fn(**kwargs)
    except ValueError:
        raise
    except Exception as exc:
        logger.exception("[IMPORT_COMPLETE] Error al guardar %s", section_name)
        raise ValueError(
            f"No se pudieron guardar {section_name}. Revisa duplicados o referencias invalidas en esa hoja."
        ) from exc


def _tag_section_errors(section_name: str, errors: list[dict]) -> list[dict]:
    tagged_errors = []
    for error in errors or []:
        if isinstance(error, dict):
            tagged_errors.append({"sheet": section_name, **error})
        else:
            tagged_errors.append({"sheet": section_name, "errors": [str(error)]})
    return tagged_errors


def _activity_import_key(lote_id, actividad, unidad, fecha, cantidad, factor_emision) -> tuple:
    return (
        lote_id,
        _normalize_text(actividad, lower=True),
        _normalize_text(unidad, lower=True),
        fecha.isoformat() if hasattr(fecha, "isoformat") else str(fecha),
        Decimal(str(cantidad)),
        Decimal(str(factor_emision)),
    )


def _confirm_complete_activities(rows: list[dict], empresa: Empresa) -> dict:
    created = 0
    omitted = 0
    duplicated = 0
    rejected = 0
    errors: list[dict] = []
    seen_keys = set()
    pending: list[EmisionLote] = []

    valid_rows = [row for row in rows if row.get("status") == "valid"]
    lote_ids = {
        (row.get("data") or {}).get("id_lote")
        for row in valid_rows
        if (row.get("data") or {}).get("id_lote")
    }
    lotes_by_id = {
        lote.id_lote: lote
        for lote in Lote.objects.select_related("empresa", "unidad_operativa").filter(id_lote__in=lote_ids)
    }
    existing_keys = {
        _activity_import_key(
            actividad.lote_id,
            actividad.actividad,
            actividad.unidad,
            actividad.fecha,
            actividad.cantidad,
            actividad.factor_emision,
        )
        for actividad in EmisionLote.objects.filter(
            lote__id_lote__in=lote_ids,
        ).only("lote_id", "actividad", "unidad", "fecha", "cantidad", "factor_emision")
    }

    for row in rows:
        row_number = row.get("row_number")
        if row.get("status") and row.get("status") != "valid":
            rejected += 1
            errors.append(
                {
                    "row_number": row_number,
                    "errors": row.get("errors") or ["fila no valida"],
                }
            )
            continue

        data = row.get("data") or {}
        lote = lotes_by_id.get(data.get("id_lote"))
        if lote is None:
            rejected += 1
            errors.append({"row_number": row_number, "errors": ["id_lote no existe"]})
            continue

        try:
            cantidad = Decimal(str(data["cantidad"]))
            factor_emision = Decimal(str(data["factor_emision"]))
            key = _activity_import_key(
                lote.pk,
                data["actividad"],
                data["unidad"],
                data["fecha"],
                cantidad,
                factor_emision,
            )
        except Exception:
            rejected += 1
            errors.append({"row_number": row_number, "errors": ["actividad invalida"]})
            continue

        if key in seen_keys or key in existing_keys:
            duplicated += 1
            omitted += 1
            continue
        seen_keys.add(key)

        actividad = EmisionLote(
            lote=lote,
            empresa=lote.empresa or empresa,
            unidad_operativa=lote.unidad_operativa,
            actividad=format_activity_display_name(data["actividad"]),
            actividad_key=normalize_activity_key(data["actividad"]),
            categoria=data.get("categoria") or "",
            cantidad=cantidad,
            unidad=data["unidad"],
            fecha=data["fecha"],
            factor_emision=factor_emision,
            emisiones_kg_co2e=cantidad * factor_emision,
            tipo_asignacion=EmisionLote.TipoAsignacion.LOTE,
        )
        pending.append(actividad)

    if pending:
        try:
            EmisionLote.objects.bulk_create(pending, batch_size=500)
            created = len(pending)
        except IntegrityError:
            logger.exception("[IMPORT_COMPLETE] Error al guardar actividades en bloque")
            for actividad in pending:
                try:
                    actividad.save()
                    created += 1
                except IntegrityError:
                    duplicated += 1
                    omitted += 1
                except Exception:
                    rejected += 1
            if rejected:
                errors.append(
                    {
                        "row_number": None,
                        "errors": ["Algunas actividades no se pudieron guardar. Revisa duplicados o referencias."],
                    }
                )

    return {
        "creados": created,
        "created": created,
        "actualizados": 0,
        "omitidos": omitted,
        "duplicados": duplicated,
        "rechazados": rejected,
        "errores": errors,
    }


def _row_processing_error(sheet_name: str, row: dict, exc: Exception) -> tuple[dict, list[str], list[str]]:
    row_number = row.get("row_number", "?") if isinstance(row, dict) else "?"
    logger.exception("[IMPORT_COMPLETE] Error al procesar hoja=%s fila=%s", sheet_name, row_number)
    return {}, [f"No se pudo procesar la fila {row_number} de la hoja {sheet_name}."], []


class ImportadorEmpresaCompleta:
    @staticmethod
    def previsualizar(uploaded_file) -> dict:
        """Preview de importación de empresa completa desde XLSX."""
        name = Path(uploaded_file.name or "").suffix.lower()
        if name != ".xlsx":
            raise ValueError("Solo se permiten archivos XLSX")

        batch_id = uuid.uuid4().hex
        blocking_errors = []
        
        # Leer todas las hojas en una sola pasada para evitar bloqueos con archivos grandes.
        sheets = _read_normalized_xlsx_sheets(
            uploaded_file,
            ["empresa", "unidades", "lotes", "actividades", "factores"],
        )
        empresa_rows = sheets["empresa"]
        unidades_rows = sheets["unidades"]
        lotes_rows = sheets["lotes"]
        actividades_rows = sheets["actividades"]
        factores_rows = sheets["factores"]

        # Procesar empresa
        empresa_data = None
        empresa_errors = []
        empresa_activa = None
        if empresa_rows:
            row = empresa_rows[0]
            try:
                empresa_data, empresa_errors = _build_company_payload(row)
            except Exception as exc:
                empresa_data, empresa_errors, _warnings = _row_processing_error("empresa", row, exc)
            if empresa_errors:
                blocking_errors.extend([f"Empresa: {e}" for e in empresa_errors])
            
            # Crear objeto empresa en memoria para usar como referencia en validaciones
            # Si la empresa no existe aún, generamos el empresa_id a partir del nombre
            if empresa_data and empresa_data.get("nombre"):
                generated_empresa_id = empresa_data.get("empresa_id") or normalize_identifier(empresa_data.get("nombre")) or "EMPRESA_GENERAL"
                # Crear objeto en memoria (sin guardar)
                empresa_activa = Empresa(
                    empresa_id=generated_empresa_id,
                    nombre=empresa_data.get("nombre"),
                    rut=empresa_data.get("rut", ""),
                    region=empresa_data.get("region", ""),
                    comuna=empresa_data.get("comuna", ""),
                    direccion=empresa_data.get("direccion", ""),
                    rubro=empresa_data.get("rubro", ""),
                    email=empresa_data.get("email", ""),
                    telefono=empresa_data.get("telefono", ""),
                    contacto=empresa_data.get("contacto", ""),
                    observaciones=empresa_data.get("observaciones", ""),
                )
                # Actualizar empresa_data con el empresa_id generado
                empresa_data["empresa_id"] = generated_empresa_id
        else:
            blocking_errors.append("La hoja empresa no contiene filas validas para importar")

        if not empresa_data or not empresa_data.get("empresa_id"):
            blocking_errors.append("No se encontro una empresa valida en la hoja empresa")

        # Procesar unidades
        unidades_preview = {
            "total": len(unidades_rows),
            "validas": 0,
            "errores": 0,
            "rows": []
        }
        unidades_by_id = {}
        unidades_by_name = {}
        for row in unidades_rows:
            try:
                data, errors, warnings = _build_unit_payload(row, empresa_activa=empresa_activa)
            except Exception as exc:
                data, errors, warnings = _row_processing_error("unidades", row, exc)
            unidad_id = data.get("unidad_id")
            nombre_unidad = _normalize_text(data.get("nombre"), lower=True)
            status = "valid" if not errors else "error"
            if status == "valid":
                unidades_preview["validas"] += 1
                unidades_by_id[unidad_id] = _strip_runtime_objects(data)
                if nombre_unidad:
                    unidades_by_name[nombre_unidad] = _strip_runtime_objects(data)
            else:
                unidades_preview["errores"] += 1
            
            unidades_preview["rows"].append({
                "row_number": row.get("row_number"),
                "status": status,
                "errors": errors,
                "warnings": warnings,
                "raw": {k: v for k, v in row.items() if k != "row_number"},
                "data": _strip_runtime_objects(data)
            })

        # Procesar factores antes de actividades para que las actividades del mismo XLSX
        # puedan usar factores que aun no existen en la base de datos.
        factores_preview = {
            "total": len(factores_rows),
            "validos": 0,
            "errores": 0,
            "rows": []
        }
        for row in factores_rows:
            try:
                data, errors = _build_row_payload(row)
            except Exception as exc:
                data, errors, _warnings = _row_processing_error("factores", row, exc)
            status = "valid" if not errors else "error"
            if status == "valid":
                factores_preview["validos"] += 1
            else:
                factores_preview["errores"] += 1

            factores_preview["rows"].append({
                "row_number": row.get("row_number"),
                "status": status,
                "errors": errors,
                "raw": {k: v for k, v in row.items() if k != "row_number"},
                "data": {**data, "factor_emision": str(data.get("factor_emision", ""))}
            })
        factor_lookup = _build_factor_lookup(factores_preview["rows"])

        # Procesar lotes
        lotes_preview = {
            "total": len(lotes_rows),
            "validos": 0,
            "errores": 0,
            "rows": []
        }
        lotes_by_id = {}
        for row in lotes_rows:
            try:
                data, errors, warnings = _build_lote_payload(row)
            except Exception as exc:
                data, errors, warnings = _row_processing_error("lotes", row, exc)
            id_lote = data.get("id_lote")
            empresa_id = data.get("empresa_id")

            if empresa_activa is not None and (not empresa_id or empresa_id == empresa_activa.empresa_id):
                _remove_error(errors, "empresa_id no existe")
                _remove_error(errors, "empresa es obligatoria")
                data["empresa_id"] = empresa_activa.empresa_id
                data["empresa_aserradero"] = data.get("empresa_aserradero") or empresa_activa.nombre
            
            # Validar que unidad existe
            unidad_id = data.get("unidad_id")
            if unidad_id in unidades_by_id:
                _remove_error(errors, "unidad_id no existe")
            elif _normalize_text(unidad_id, lower=True) in unidades_by_name:
                unidad_data = unidades_by_name[_normalize_text(unidad_id, lower=True)]
                _remove_error(errors, "unidad_id no existe")
                data["unidad_id"] = unidad_data.get("unidad_id")
                data["empresa_id"] = unidad_data.get("empresa_id") or data.get("empresa_id")
            elif unidad_id and not UnidadOperativa.objects.filter(unidad_id=unidad_id).exists():
                errors.append(f"unidad_id {unidad_id} no existe en el archivo ni en la base de datos")
            
            status = "valid" if not errors else "error"
            if status == "valid":
                lotes_preview["validos"] += 1
                lotes_by_id[id_lote] = _strip_runtime_objects(data)
            else:
                lotes_preview["errores"] += 1
            
            lotes_preview["rows"].append({
                "row_number": row.get("row_number"),
                "status": status,
                "errors": errors,
                "warnings": warnings,
                "raw": {k: v for k, v in row.items() if k != "row_number"},
                "data": _strip_runtime_objects(data)
            })

        # Procesar actividades
        actividades_preview = {
            "total": len(actividades_rows),
            "validas": 0,
            "errores": 0,
            "factores_encontrados": 0,
            "factores_faltantes": 0,
            "rows": []
        }
        for row in actividades_rows:
            try:
                data, errors, warnings = _build_complete_activity_payload(row, empresa_activa=empresa_activa)
            except Exception as exc:
                data, errors, warnings = _row_processing_error("actividades", row, exc)
            
            # Validar que lote existe si está referenciado
            id_lote = data.get("id_lote")
            unidad_id = data.get("unidad_id")
            empresa_id = data.get("empresa_id")

            if id_lote in lotes_by_id:
                lote_data = lotes_by_id[id_lote]
                _remove_error(errors, "id_lote no existe")
                data["empresa_id"] = lote_data.get("empresa_id") or data.get("empresa_id")
                data["unidad_id"] = lote_data.get("unidad_id") or data.get("unidad_id")
                data["tipo_asignacion"] = EmisionLote.TipoAsignacion.LOTE
            elif (
                id_lote
                and not Lote.objects.filter(id_lote=id_lote).exists()
                and f"id_lote {id_lote} no existe" not in errors
            ):
                errors.append(f"id_lote {id_lote} no existe")

            if unidad_id in unidades_by_id:
                unidad_data = unidades_by_id[unidad_id]
                _remove_error(errors, "unidad_id no existe")
                data["empresa_id"] = unidad_data.get("empresa_id") or data.get("empresa_id")
                if not data.get("id_lote"):
                    data["tipo_asignacion"] = EmisionLote.TipoAsignacion.UNIDAD
            elif _normalize_text(unidad_id, lower=True) in unidades_by_name:
                unidad_data = unidades_by_name[_normalize_text(unidad_id, lower=True)]
                _remove_error(errors, "unidad_id no existe")
                data["unidad_id"] = unidad_data.get("unidad_id")
                data["empresa_id"] = unidad_data.get("empresa_id") or data.get("empresa_id")
                if not data.get("id_lote"):
                    data["tipo_asignacion"] = EmisionLote.TipoAsignacion.UNIDAD

            if empresa_activa is not None and (not empresa_id or empresa_id == empresa_activa.empresa_id):
                _remove_error(errors, "empresa_id no existe")
                data["empresa_id"] = empresa_activa.empresa_id

            _apply_file_factor(data, errors, factor_lookup)
            if data.get("actividad") and data.get("unidad") and not data.get("factor_emision"):
                errors.append("factor de emision no encontrado")
            
            if data.get("factor_emision"):
                actividades_preview["factores_encontrados"] += 1
            else:
                actividades_preview["factores_faltantes"] += 1
            
            status = "valid" if not errors else "error"
            if status == "valid":
                actividades_preview["validas"] += 1
            else:
                actividades_preview["errores"] += 1
            
            actividades_preview["rows"].append({
                "row_number": row.get("row_number"),
                "status": status,
                "errors": errors,
                "warnings": warnings,
                "raw": {k: v for k, v in row.items() if k != "row_number"},
                "data": _strip_runtime_objects(data)
            })

        # Cache el resultado
        cache_data = {
            "empresa": {
                "data": empresa_data or {},
                "errors": empresa_errors,
                "status": "valid" if not empresa_errors else "error"
            },
            "unidades": unidades_preview,
            "lotes": lotes_preview,
            "actividades": actividades_preview,
            "factores": factores_preview,
            "blocking_errors": blocking_errors,
        }
        
        cache.set(
            f"{COMPLETE_IMPORT_CACHE_PREFIX}{batch_id}",
            cache_data,
            timeout=FACTOR_IMPORT_CACHE_TTL_SECONDS,
        )

        return {
            "batch_id": batch_id,
            "empresa": cache_data["empresa"],
            "unidades": unidades_preview,
            "lotes": lotes_preview,
            "actividades": actividades_preview,
            "factores": factores_preview,
            "blocking_errors": blocking_errors,
        }

    @staticmethod
    def confirmar(batch_id: str) -> dict:
        """Confirma la importación completa."""
        logger.info(f"[IMPORT_COMPLETE] Iniciando confirmación para batch_id={batch_id}")
        
        cached = cache.get(f"{COMPLETE_IMPORT_CACHE_PREFIX}{batch_id}")
        if not cached:
            logger.error(f"[IMPORT_COMPLETE] batch_id no existe o expiró: {batch_id}")
            raise ValueError("El batch_id no existe o expiró")

        logger.info(f"[IMPORT_COMPLETE] Cache data recuperado: empresa_data={bool(cached.get('empresa'))}, unidades={len(cached.get('unidades', {}).get('rows', []))}, lotes={len(cached.get('lotes', {}).get('rows', []))}, actividades={len(cached.get('actividades', {}).get('rows', []))}, factores={len(cached.get('factores', {}).get('rows', []))}")

        empresa_data = cached.get("empresa", {}).get("data", {})
        if not empresa_data or not empresa_data.get("empresa_id"):
            logger.error(f"[IMPORT_COMPLETE] No contiene empresa válida para guardar")
            raise ValueError("La importación no contiene una empresa valida para guardar")

        company_payload, company_errors = _build_company_payload(empresa_data)
        if company_errors:
            logger.error(f"[IMPORT_COMPLETE] Errores al validar empresa: {company_errors}")
            raise ValueError("; ".join(company_errors))

        company_rows = [
            {
                "row_number": 1,
                "status": "valid",
                "data": company_payload,
            }
        ]

        logger.info(f"[IMPORT_COMPLETE] Guardando empresa: empresa_id={company_payload.get('empresa_id')}")
        company_summary = _confirm_section(
            "empresa",
            ImportadorEmpresas.confirmar,
            rows=company_rows,
        )
        logger.info(f"[IMPORT_COMPLETE] Resultado empresa: {company_summary}")
        
        if company_summary.get("rechazados"):
            errores = company_summary.get("errores") or []
            logger.error(f"[IMPORT_COMPLETE] Empresa rechazada: {errores}")
            raise ValueError("No se pudo guardar la empresa: " + "; ".join(
                ", ".join(err.get("errors", [])) if isinstance(err, dict) else str(err)
                for err in errores
            ))

        created_empresa = Empresa.objects.get(empresa_id=company_payload["empresa_id"])
        logger.info(f"[IMPORT_COMPLETE] Empresa creada/actualizada: {created_empresa.empresa_id}")

        # Guardar factores primero
        logger.info(f"[IMPORT_COMPLETE] Procesando {len(cached['factores'].get('rows', []))} factores")
        factores_summary = _confirm_section(
            "factores",
            ImportadorFactores.confirmar,
            rows=cached["factores"].get("rows", []),
        )
        logger.info(f"[IMPORT_COMPLETE] Resultado factores: creados={factores_summary.get('creados', 0)}, errores={len(factores_summary.get('errores', []))}")
        if factores_summary.get("errores"):
            logger.warning(f"[IMPORT_COMPLETE] Errores en factores: {factores_summary.get('errores')}")

        # Guardar unidades
        logger.info(f"[IMPORT_COMPLETE] Procesando {len(cached['unidades'].get('rows', []))} unidades")
        unidades_summary = _confirm_section(
            "unidades",
            ImportadorUnidadesOperativas.confirmar,
            rows=cached["unidades"].get("rows", []),
            empresa_activa=created_empresa,
        )
        logger.info(f"[IMPORT_COMPLETE] Resultado unidades: creados={unidades_summary.get('creados', 0)}, errores={len(unidades_summary.get('errores', []))}")
        if unidades_summary.get("errores"):
            logger.warning(f"[IMPORT_COMPLETE] Errores en unidades: {unidades_summary.get('errores')}")

        # Guardar lotes
        logger.info(f"[IMPORT_COMPLETE] Procesando {len(cached['lotes'].get('rows', []))} lotes")
        lotes_summary = _confirm_section(
            "lotes",
            ImportadorLotes.confirmar,
            rows=cached["lotes"].get("rows", []),
            empresa_activa=created_empresa,
        )
        logger.info(f"[IMPORT_COMPLETE] Resultado lotes: creados={lotes_summary.get('creados', 0)}, rechazados={lotes_summary.get('rechazados', 0)}, errores={len(lotes_summary.get('errores', []))}")
        if lotes_summary.get("errores"):
            logger.warning(f"[IMPORT_COMPLETE] Errores en lotes: {lotes_summary.get('errores')}")

        # Guardar actividades
        logger.info(f"[IMPORT_COMPLETE] Procesando {len(cached['actividades'].get('rows', []))} actividades")
        actividades_summary = _confirm_complete_activities(
            cached["actividades"].get("rows", []),
            created_empresa,
        )
        logger.info(f"[IMPORT_COMPLETE] Resultado actividades: creados={actividades_summary.get('creados', 0)}, rechazados={actividades_summary.get('rechazados', 0)}, errores={len(actividades_summary.get('errores', []))}")
        if actividades_summary.get("errores"):
            logger.warning(f"[IMPORT_COMPLETE] Errores en actividades: {actividades_summary.get('errores')}")

        resultado = {
            "creados": 1,
            "empresa_id": created_empresa.empresa_id,
            "empresa_nombre": created_empresa.nombre,
            "unidades_creadas": unidades_summary.get("creados", 0),
            "lotes_creados": lotes_summary.get("creados", 0),
            "actividades_creadas": actividades_summary.get("creados", 0),
            "factores_creados": factores_summary.get("creados", 0),
            "errores": [
                *_tag_section_errors("unidades", unidades_summary.get("errores", [])),
                *_tag_section_errors("lotes", lotes_summary.get("errores", [])),
                *_tag_section_errors("actividades", actividades_summary.get("errores", [])),
                *_tag_section_errors("factores", factores_summary.get("errores", [])),
            ],
        }
        
        logger.info(f"[IMPORT_COMPLETE] Importación completada: {resultado}")
        return resultado
