"""Generador de plantillas XLSX descargables para importación de datos."""

from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _add_header_style(ws, row_num, headers):
    """Aplica estilos a la fila de encabezados."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _add_example_style(ws, row_num):
    """Aplica estilos a la fila de ejemplo."""
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    example_font = Font(italic=True, size=10)
    example_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = example_alignment
        cell.border = thin_border


def _add_instructions_sheet(workbook):
    """Agrega una hoja de instrucciones a la plantilla."""
    ws = workbook.create_sheet("Instrucciones", 0)
    
    instructions = [
        ("PLANTILLA DE IMPORTACIÓN - HUELLA DE CARBONO", None),
        ("", None),
        ("Esta plantilla permite importar datos completos de una empresa (empresa, unidades, lotes, actividades y factores de emisión).", None),
        ("", None),
        ("HOJAS INCLUIDAS:", None),
        ("1. Empresa - Datos básicos de la empresa", None),
        ("2. Unidades - Unidades operativas (aserradero, patio, etc.)", None),
        ("3. Lotes - Lotes de madera procesados", None),
        ("4. Actividades - Actividades realizadas en lotes con factores de emisión", None),
        ("5. Factores - Factores de emisión personalizados (opcional)", None),
        ("", None),
        ("INSTRUCCIONES GENERALES:", None),
        ("• Completa todas las celdas marcadas como [REQUERIDO]", None),
        ("• Usa exactamente los formatos indicados (ej: YYYY-MM-DD para fechas)", None),
        ("• No modifiques los nombres de las hojas ni los encabezados", None),
        ("• Los tipos de unidades deben ser uno de: Fundo Forestal, Transporte, Aserradero, Acopio, Secado, Administración, Bodega, Planta Industrial", None),
        ("• Las especias de madera comunes: Pino Radiata, Eucalipto, Álamo, Raulí, Roble", None),
        ("• Fecha formato: YYYY-MM-DD (ejemplo: 2025-01-15)", None),
        ("", None),
        ("VALIDACIÓN:", None),
        ("• El sistema valida automáticamente los datos antes de confirmar", None),
        ("• Errores se muestran con explicaciones claras y ejemplos", None),
        ("• Advertencias se indican pero no bloquean la importación", None),
    ]
    
    for row_num, (text, _) in enumerate(instructions, 1):
        ws.append([text])
        cell = ws.cell(row=row_num, column=1)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws.column_dimensions['A'].width = 100
    ws.row_dimensions[1].font = Font(bold=True, size=14)


def generate_complete_import_template() -> BytesIO:
    """Genera una plantilla XLSX descargable para importación completa de empresa."""
    workbook = Workbook()
    
    # Quitar hoja por defecto
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    # Agregar hoja de instrucciones
    _add_instructions_sheet(workbook)
    
    # ===== EMPRESA =====
    empresa_sheet = workbook.create_sheet("empresa")
    empresa_headers = [
        "ID Empresa [REQUERIDO]",
        "Nombre [REQUERIDO]",
        "RUT",
        "Región",
        "Comuna",
        "Dirección",
        "Rubro",
        "Email",
        "Teléfono",
        "Contacto",
        "Observaciones"
    ]
    empresa_sheet.append(empresa_headers)
    _add_header_style(empresa_sheet, 1, empresa_headers)
    
    # Ejemplo empresa
    empresa_sheet.append([
        "EMP-ASERRADERO-001",
        "Aserradero Los Andes",
        "76.543.210-9",
        "La Araucanía",
        "Temuco",
        "Ruta 5 Sur km 123",
        "Forestal",
        "contacto@aserradero.cl",
        "+56 9 1234 5678",
        "Juan Pérez",
        "Importación inicial"
    ])
    _add_example_style(empresa_sheet, 2)
    
    # Ajustar ancho de columnas
    for col in range(1, len(empresa_headers) + 1):
        empresa_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== UNIDADES =====
    unidades_sheet = workbook.create_sheet("unidades")
    unidades_headers = [
        "ID Unidad [REQUERIDO]",
        "Nombre [REQUERIDO]",
        "Tipo [REQUERIDO]",
        "Región",
        "Comuna",
        "Dirección"
    ]
    unidades_sheet.append(unidades_headers)
    _add_header_style(unidades_sheet, 1, unidades_headers)
    
    # Ejemplo unidad
    unidades_sheet.append([
        "UNI-ASERR-001",
        "Aserradero Principal",
        "Aserradero",
        "La Araucanía",
        "Temuco",
        "Ruta 5 Sur km 123"
    ])
    _add_example_style(unidades_sheet, 2)
    
    # Ejemplo adicional (patio/acopio)
    unidades_sheet.append([
        "UNI-PATIO-001",
        "Patio de Acopio",
        "Acopio",
        "La Araucanía",
        "Temuco",
        "Ruta 5 Sur km 125"
    ])
    _add_example_style(unidades_sheet, 3)
    
    # Ajustar ancho
    for col in range(1, len(unidades_headers) + 1):
        unidades_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== LOTES =====
    lotes_sheet = workbook.create_sheet("lotes")
    lotes_headers = [
        "ID Lote [REQUERIDO]",
        "ID Unidad [REQUERIDO]",
        "Fecha [REQUERIDO]",
        "Especie [REQUERIDO]",
        "Volumen (m³) [REQUERIDO]",
        "Origen [REQUERIDO]"
    ]
    lotes_sheet.append(lotes_headers)
    _add_header_style(lotes_sheet, 1, lotes_headers)
    
    # Ejemplo lote
    lotes_sheet.append([
        "LOT-2025-001",
        "UNI-ASERR-001",
        "2025-01-15",
        "Pino Radiata",
        250.5,
        "Cosecha sector norte, fundo Las Vertientes"
    ])
    _add_example_style(lotes_sheet, 2)
    
    # Ajustar ancho
    for col in range(1, len(lotes_headers) + 1):
        lotes_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== ACTIVIDADES =====
    actividades_sheet = workbook.create_sheet("actividades")
    actividades_headers = [
        "ID Unidad [REQUERIDO]",
        "ID Lote [REQUERIDO]",
        "Actividad [REQUERIDO]",
        "Cantidad [REQUERIDO]",
        "Unidad [REQUERIDO]",
        "Fecha [REQUERIDO]"
    ]
    actividades_sheet.append(actividades_headers)
    _add_header_style(actividades_sheet, 1, actividades_headers)
    
    # Ejemplo actividad
    actividades_sheet.append([
        "UNI-ASERR-001",
        "LOT-2025-001",
        "Diesel",
        50,
        "litros",
        "2025-01-15"
    ])
    _add_example_style(actividades_sheet, 2)
    
    # Segundo ejemplo
    actividades_sheet.append([
        "UNI-ASERR-001",
        "LOT-2025-001",
        "Electricidad",
        150,
        "kWh",
        "2025-01-15"
    ])
    _add_example_style(actividades_sheet, 3)
    
    # Ajustar ancho
    for col in range(1, len(actividades_headers) + 1):
        actividades_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== FACTORES =====
    factores_sheet = workbook.create_sheet("factores")
    factores_headers = [
        "Actividad [REQUERIDO]",
        "Unidad [REQUERIDO]",
        "Factor de Emisión [REQUERIDO]",
        "Fuente",
        "Año"
    ]
    factores_sheet.append(factores_headers)
    _add_header_style(factores_sheet, 1, factores_headers)
    
    # Ejemplo factores
    factores_sheet.append([
        "Diesel",
        "litros",
        2.68,
        "Base de datos nacional",
        2024
    ])
    _add_example_style(factores_sheet, 2)
    
    factores_sheet.append([
        "Electricidad",
        "kWh",
        0.392,
        "IPCC 2023",
        2024
    ])
    _add_example_style(factores_sheet, 3)
    
    # Ajustar ancho
    for col in range(1, len(factores_headers) + 1):
        factores_sheet.column_dimensions[get_column_letter(col)].width = 25
    
    # Guardar en BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output
