import hashlib
import json
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from .services import sanitize


@dataclass(frozen=True)
class ParsedTabularResource:
    sheets: list
    headers: list
    rows_read: int
    rows: list
    rejected: list


class TabularResourceParser:
    def parse(self, path, **context):
        raise NotImplementedError

    def read_xlsx(self, path):
        if not zipfile.is_zipfile(path):
            raise ValueError("El recurso no es un archivo XLSX válido.")
        try:
            workbook=load_workbook(path,read_only=True,data_only=True)
        except Exception as exc:
            raise ValueError("No fue posible abrir el workbook XLSX.") from exc
        if not workbook.sheetnames:
            raise ValueError("El workbook XLSX no contiene hojas.")
        worksheet=workbook[workbook.sheetnames[0]]
        iterator=worksheet.iter_rows(values_only=True)
        try: headers=[str(value).strip() if value is not None else "" for value in next(iterator)]
        except StopIteration as exc: raise ValueError("El workbook XLSX está vacío.") from exc
        return workbook,worksheet,headers,iterator


class RetcHazardousWasteParser(TabularResourceParser):
    HEADERS=("año","id_vu","id_rol_establecimiento","rol_establecimiento","rut_razon_social","razon_social","ciiu6_id","ciiu6","ciiu4_id","ciiu4","rubro","rubro_id","codigo_unico_territorial","comuna","provincia","region","latitud","longitud","cantidad_kilos","cantidad_toneladas","id_contaminantes","contaminantes","id_peligrosidad","peligrosidad","id_lista_a","lista_a","id_estado_materia","estado_materia")
    INTEGER_FIELDS=("año","id_vu","id_rol_establecimiento","rubro_id","codigo_unico_territorial","id_estado_materia")
    DECIMAL_FIELDS=("cantidad_kilos","cantidad_toneladas")

    def parse(self,path,year,**context):
        workbook,worksheet,headers,iterator=self.read_xlsx(path)
        if tuple(headers)!=self.HEADERS:
            iterator.close();workbook.close();raise ValueError("El esquema XLSX RETC no coincide con el contrato conocido.")
        rows=[];rejected=[];rows_read=0
        try:
            for row_number,values in enumerate(iterator,start=2):
                rows_read+=1
                if all(value is None or str(value).strip()=="" for value in values):
                    rejected.append({"row_number":row_number,"reason":"fila_vacia"});continue
                raw=sanitize({header:value for header,value in zip(headers,values)})
                try: normalized=self._normalize(raw,year)
                except (ValueError,TypeError,InvalidOperation) as exc:
                    rejected.append({"row_number":row_number,"reason":str(exc)});continue
                encoded=json.dumps(raw,sort_keys=True,separators=(",",":"),ensure_ascii=False,default=str).encode()
                rows.append({"source_row_number":row_number,"row_hash":hashlib.sha256(encoded).hexdigest(),"raw_row":raw,**normalized})
        finally: iterator.close();workbook.close()
        return ParsedTabularResource([worksheet.title],headers,rows_read,rows,rejected)

    def _normalize(self,raw,expected_year):
        normalized={}
        for field in self.INTEGER_FIELDS:
            if raw.get(field) is None: raise ValueError(f"campo_obligatorio_ausente:{field}")
            normalized[field if field!="año" else "year"]=int(raw[field])
        if normalized["year"]!=expected_year: raise ValueError("año_fuera_del_recurso")
        for field in self.DECIMAL_FIELDS:
            normalized[field]=Decimal(str(raw[field])) if raw.get(field) is not None else None
        if normalized["cantidad_toneladas"] is None: raise ValueError("campo_obligatorio_ausente:cantidad_toneladas")
        text_fields=("rol_establecimiento","rut_razon_social","razon_social","ciiu6_id","ciiu6","ciiu4_id","ciiu4","rubro","comuna","provincia","region","id_contaminantes","contaminantes","id_peligrosidad","peligrosidad","id_lista_a","lista_a","estado_materia")
        for field in text_fields: normalized[field]="" if raw.get(field) is None else str(raw[field]).strip()
        for field in ("latitud","longitud"): normalized[f"{field}_raw"]="" if raw.get(field) is None else str(raw[field])
        return normalized


class HuellaChileEmissionFactorParser(TabularResourceParser):
    SHEET="RESUMEN"
    HEADERS=("Alcance","Categoría","Subcategoría","Nombre","Auxiliar","Unidad del dato de actividad","Factor de emisión","Unidades del factor de emisión","Fuente 1","Fuente 2","Fuente 3")

    def parse(self,path,year,**context):
        if not zipfile.is_zipfile(path): raise ValueError("El recurso no es un archivo XLSX válido.")
        try:
            formulas=load_workbook(path,read_only=False,data_only=False)
            cached=load_workbook(path,read_only=False,data_only=True)
        except Exception as exc: raise ValueError("No fue posible abrir el workbook HuellaChile.") from exc
        try:
            if self.SHEET not in formulas.sheetnames: raise ValueError("El workbook no contiene la hoja RESUMEN.")
            formula_sheet=formulas[self.SHEET];cached_sheet=cached[self.SHEET]
            headers=tuple(cached_sheet.cell(8,column).value for column in range(2,13))
            if headers!=self.HEADERS: raise ValueError("El esquema RESUMEN de HuellaChile no coincide con el contrato conocido.")
            rows=[];rejected=[];rows_read=0
            for row_number in range(9,cached_sheet.max_row+1):
                values=[cached_sheet.cell(row_number,column).value for column in range(2,13)]
                if all(value is None for value in values): continue
                rows_read+=1
                if any(values[index] is None for index in range(8)):
                    rejected.append({"row_number":row_number,"reason":"campos_resumen_obligatorios_ausentes"});continue
                factor_cell=formula_sheet.cell(row_number,8);formula_value=factor_cell.value
                formula_text=formula_value.text if isinstance(formula_value,ArrayFormula) else str(formula_value) if factor_cell.data_type=="f" else ""
                published=values[6];numeric=isinstance(published,(int,float,Decimal)) and not isinstance(published,bool)
                raw={}
                for offset,header in enumerate(self.HEADERS,start=2):
                    formula_cell=formula_sheet.cell(row_number,offset);cached_value=cached_sheet.cell(row_number,offset).value
                    if formula_cell.data_type=="f" or isinstance(formula_cell.value,ArrayFormula):
                        original=formula_cell.value.text if isinstance(formula_cell.value,ArrayFormula) else str(formula_cell.value)
                        raw[header]={"formula":original,"cached":cached_value}
                    else: raw[header]=cached_value
                encoded=json.dumps(raw,sort_keys=True,separators=(",",":"),ensure_ascii=False,default=str).encode()
                rows.append({"sheet_name":self.SHEET,"source_row_number":row_number,"row_hash":hashlib.sha256(encoded).hexdigest(),"raw_row":raw,"dataset_year":year,"alcance":str(values[0]),"categoria":str(values[1]),"subcategoria":str(values[2]),"actividad":str(values[3]),"auxiliar":str(values[4]),"unidad_actividad":str(values[5]),"factor_value":Decimal(str(published)) if numeric else None,"published_value_raw":str(published),"unidad_factor":str(values[7]),"technical_source_1":"" if values[8] is None else str(values[8]),"technical_source_2":"" if values[9] is None else str(values[9]),"technical_source_3":"" if values[10] is None else str(values[10]),"formula_original":formula_text,"cached_value_available":numeric})
            sheet_metadata=[{"name":sheet.title,"state":sheet.sheet_state,"rows":sheet.max_row,"columns":sheet.max_column,"merged_cells":len(sheet.merged_cells.ranges)} for sheet in formulas.worksheets]
            references=sorted({str(value) for row in rows for value in (row["technical_source_1"],row["technical_source_2"],row["technical_source_3"]) if value and value!="-"})
            metadata={"sheets":sheet_metadata,"hidden_sheets":[sheet["name"] for sheet in sheet_metadata if sheet["state"]!="visible"],"headers":list(self.HEADERS),"references":references,"workbook":{"creator":formulas.properties.creator,"title":formulas.properties.title,"created":str(formulas.properties.created or ""),"modified":str(formulas.properties.modified or ""),"last_modified_by":formulas.properties.lastModifiedBy}}
            return ParsedTabularResource([sheet["name"] for sheet in sheet_metadata],list(self.HEADERS),rows_read,rows,rejected),metadata
        finally: formulas.close();cached.close()
